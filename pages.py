"""
pages.py — 所有頁面
整合功能：智慧網格導航(Auto-Scroll)、小數點淨化、月詳細介面優化、無語法縮寫安全版。
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import data_manager as dm
import ui_utils as U

_font_family = getattr(U, 'F', 'Microsoft JhengHei')
U.FM = (_font_family, 11)        
U.FMB = (_font_family, 11, 'bold') 
U.FS = (_font_family, 10)

# 💡 安全的全域樣式設定器 (避免提早觸發空白 tk 視窗)
_global_style_applied = False
def _apply_global_style(widget):
    global _global_style_applied
    if not _global_style_applied:
        try:
            root = widget.winfo_toplevel()
            root.option_add("*TCombobox*Listbox.font", U.FM)
            root.option_add("*Listbox.font", U.FM)
            style = ttk.Style()
            style.configure("TCombobox", font=U.FM)
            style.configure("TButton", font=U.FM)
            _global_style_applied = True
        except Exception:
            pass

# 💡 共用小工具：小數點淨化 (90.0 -> 90)
def _fmt(v):
    if v == '' or v is None: 
        return ''
    try:
        f = float(v)
        return str(int(f)) if f.is_integer() else str(f)
    except Exception:
        return str(v)

def _clean_val(v):
    if v == '' or v is None: 
        return ''
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except Exception: 
        return v

# 💡 底層輔助：安全刪除單筆月成績紀錄
def _delete_monthly_rec(data, cls, zh, ym, r_type, r_date, r_range, r_item):
    try:
        recs = data.get("monthly_records", {}).get(cls, {}).get(ym, {}).get(zh, [])
        for i, r in enumerate(recs):
            if r.get("type") == r_type and r.get("date") == r_date and r.get("range") == r_range and r.get("item") == r_item:
                del recs[i]
                return True
    except Exception: pass

    try:
        recs = data.get("monthly", {}).get(cls, {}).get(zh, {}).get(ym, [])
        for i, r in enumerate(recs):
            if r.get("type") == r_type and r.get("date") == r_date and r.get("range") == r_range and r.get("item") == r_item:
                del recs[i]
                return True
    except Exception: pass

    try:
        for stu in data.get("classes", {}).get(cls, {}).get("students", []):
            if stu.get("zh") == zh:
                for i, r in enumerate(stu.get("records", [])):
                    if r.get("type") == r_type and r.get("date") == r_date and r.get("range") == r_range and r.get("item") == r_item:
                        del stu["records"][i]
                        return True
    except Exception: pass
    return False

# 💡 共用小工具：自動綁定 Enter / 上下鍵導航，並支援智慧滾輪跟隨
def apply_matrix_nav(matrix, canvas=None):
    def focus_and_scroll(entry):
        entry.focus_set()
        entry.select_range(0, tk.END)
        # 💡 智慧追蹤：自動計算並捲動到游標位置
        if canvas:
            canvas.update_idletasks()
            try:
                inner = canvas.winfo_children()[0]
                y_offset = 0
                w = entry
                # 往上尋找相對於 inner 的 Y 座標
                while w and w != inner:
                    y_offset += w.winfo_y()
                    w = w.master
                
                i_h = inner.winfo_height()
                c_h = canvas.winfo_height()
                if i_h == 0 or c_h == 0: 
                    return
                
                top_f, bot_f = canvas.yview()
                w_h = entry.winfo_height()
                w_top_f = y_offset / i_h
                w_bot_f = (y_offset + w_h) / i_h
                
                pad_f = 40 / i_h # 上下預留舒適空間
                
                if w_bot_f > bot_f:
                    canvas.yview_moveto(w_bot_f - (c_h / i_h) + pad_f)
                elif w_top_f < top_f:
                    canvas.yview_moveto(max(0.0, w_top_f - pad_f))
            except Exception:
                pass

    for c in range(len(matrix)):
        for r in range(len(matrix[c])):
            entry = matrix[c][r]
            def make_down(curr_r, curr_c):
                def handler(event):
                    nr, nc = curr_r + 1, curr_c
                    if nr >= len(matrix[nc]):
                        nr, nc = 0, nc + 1
                    if nc < len(matrix):
                        focus_and_scroll(matrix[nc][nr])
                    return "break"
                return handler
            def make_up(curr_r, curr_c):
                def handler(event):
                    nr, nc = curr_r - 1, curr_c
                    if nr < 0:
                        nc -= 1
                        if nc >= 0: 
                            nr = len(matrix[nc]) - 1
                    if nc >= 0:
                        focus_and_scroll(matrix[nc][nr])
                    return "break"
                return handler
            
            entry.bind("<Return>", make_down(r, c))
            entry.bind("<Down>", make_down(r, c))
            entry.bind("<Up>", make_up(r, c))

def _make_score_validator(widget_root):
    def _validate(P):
        if P == '': 
            return True
        try: 
            return 0 <= float(P) <= 100
        except ValueError: 
            return False
    return (widget_root.register(_validate), '%P')

def _make_int_score_validator(widget_root):
    def _validate(P):
        if P == '': 
            return True
        if not P.isdigit(): 
            return False
        return 0 <= int(P) <= 100
    return (widget_root.register(_validate), '%P')

# 版本號（每次更新前請修改）
VERSION = "2026.01.01"

# ══════════════════════════════════════════════════════════════════════════════
# 兒美成績輸入 (原總覽)
# ══════════════════════════════════════════════════════════════════════════════
class OverviewPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._cls_var = tk.StringVar()
        # 年份預設：從 grades.json 最新資料取，否則用今年
        def _latest_year(data):
            years = []
            for cdata in data.get("classes", {}).values():
                for lt in cdata.get("leveltests", []):
                    d = lt.get("date", "")
                    if d: years.append(d[:4].replace("/",""))
                for s in cdata.get("students", []):
                    for ym in s.get("scores", {}).keys() if isinstance(s, dict) else []:
                        if ym: years.append(ym[:4])
            # 從 comments keys 取月份
            for cls_months in data.get("comments", {}).values():
                for ym in cls_months.keys():
                    if ym: years.append(ym[:4])
            return max(years) if years else str(datetime.now().year)
        self._yr_var  = tk.StringVar(value=_latest_year(self.app.data))
        self._mo_var  = tk.StringVar(value=f"{datetime.now().month:02d}")
        self._build()

    def _build(self):
        self._topbar = tk.Frame(self, bg=U.C["bg"])
        self._topbar.pack(fill="x", padx=20, pady=(14, 6))
        
        U.lbl(self._topbar, "兒美成績輸入與管理", size=18, bold=True).pack(side="left")
        
        U.btn(self._topbar, "新增成績紀錄", lambda: self.app._show("monthly"), size=12).pack(side="right", padx=10)
        U.ghost_btn(self._topbar, "📥 匯入 Excel 總表", self._import_excel, size=12).pack(side="right", padx=4)
        U.ghost_btn(self._topbar, "🔗 複製評語連結", self._update_html_form, size=12).pack(side="right", padx=4)
        
        self._book_lbl = tk.Label(self._topbar, text="", font=U.FM, bg=U.C["ok_bg"], fg=U.C["ok_fg"], padx=10, pady=3, relief="flat")
        self._book_lbl.pack(side="left", padx=10)

        frow = tk.Frame(self, bg=U.C["bg"])
        frow.pack(fill="x", padx=20, pady=(0, 8))
        U.lbl(frow, "班級", size=13, bg=U.C["bg"]).pack(side="left")
        self._cls_combo = U.combo(frow, self._cls_var, [], width=8)
        self._cls_combo.pack(side="left", padx=(4, 14))
        self._cls_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_table())
        U.lbl(frow, "年份", size=13, bg=U.C["bg"]).pack(side="left")
        tk.Spinbox(frow, textvariable=self._yr_var, from_=2016, to=2036,
                   width=6, font=U.FM, relief="flat", state="readonly",
                   highlightbackground=U.C["border"], highlightthickness=1,
                   buttonbackground=U.C["bg"]).pack(side="left", padx=(4, 10))
        U.lbl(frow, "月份", size=13, bg=U.C["bg"]).pack(side="left")
        U.combo(frow, self._mo_var, [f"{m:02d}" for m in range(1, 13)], width=5).pack(side="left", padx=(4, 10))
        U.btn(frow, "查詢", self._refresh_table, size=12).pack(side="left")

        leg = tk.Frame(self, bg=U.C["bg"])
        leg.pack(fill="x", padx=20, pady=(0, 4))
        U.lbl(leg, "💡 點擊可直接快速編輯", size=13, bg=U.C["bg"], fg=U.C["primary"]).pack(side="left", padx=(0, 15))
        for dot_color, text in [(U.C["ok_fg"], "達標"), (U.C["ng_fg"], "未達標/缺考")]:
            tk.Label(leg, text="●", fg=dot_color, bg=U.C["bg"], font=U.FS).pack(side="left")
            U.lbl(leg, text + "　", size=13, bg=U.C["bg"], fg=U.C["text_lt"]).pack(side="left")

        outer = tk.Frame(self, bg=U.C["bg"])
        outer.pack(fill="both", expand=True, padx=20, pady=(0, 6))
        self._canvas = tk.Canvas(outer, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        xsb = ttk.Scrollbar(outer, orient="horizontal", command=self._canvas.xview)
        ysb = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(xscrollcommand=xsb.set, yscrollcommand=ysb.set)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._table_frame = tk.Frame(self._canvas, bg=U.C["white"])
        self._canvas.create_window((0, 0), window=self._table_frame, anchor="nw")
        self._table_frame.bind("<Configure>", lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        U.bind_mousewheel(self._canvas)

        self._bottom = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        self._bottom.pack(fill="x", padx=20, pady=(0, 10))
        self._missing_lbl = U.lbl(self._bottom, "", size=11, bg=U.C["white"], fg=U.C["ng_fg"])
        self._missing_lbl.pack(anchor="w", padx=12, pady=6)


    def _import_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("缺少套件", "請先安裝 openpyxl 套件：\n請關閉程式後，在命令提示字元(cmd)輸入：\npip install openpyxl", parent=self)
            return

        file_path = filedialog.askopenfilename(
            title="選擇要匯入的 兒美成績輸入表 (Excel)",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path: return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            success_count = 0
            month_dict = {'一':1, '二':2, '三':3, '四':4, '五':5, '六':6, '七':7, '八':8, '九':9, '十':10, '十一':11, '十二':12}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                last_month = self._mo_var.get()
                last_cls = self._cls_var.get()
                
                r = 0
                while r < len(rows):
                    row = rows[r]
                    if len(row) > 5 and row[5] is not None:
                        label = str(row[5]).replace(" ", "").strip()
                        if label == "日期":
                            if r + 4 >= len(rows): break 
                            
                            m_val = row[0]
                            if m_val is not None and str(m_val).strip(): last_month = str(m_val).strip()
                            c_val = row[1]
                            if c_val is not None and str(c_val).strip(): last_cls = str(c_val).strip()
                                
                            stu_raw = row[4]
                            if not stu_raw: 
                                r += 5; continue
                                
                            # 💡 修正 1：精準解析學生中英文姓名
                            stu_str = str(stu_raw).strip()
                            if '(' in stu_str and ')' in stu_str:
                                zh_name = stu_str[:stu_str.index('(')].strip()
                                en_name = stu_str[stu_str.index('(')+1:stu_str.index(')')].strip()
                            elif '（' in stu_str and '）' in stu_str:
                                zh_name = stu_str[:stu_str.index('（')].strip()
                                en_name = stu_str[stu_str.index('（')+1:stu_str.index('）')].strip()
                            else:
                                zh_name = stu_str
                                en_name = ""
                                
                            # 💡 修正 2：如果系統內沒有這個班級或學生，自動幫忙建立（成績才不會隱形）
                            if last_cls not in self.app.data.get("classes", {}):
                                dm.add_class(self.app.data, last_cls)
                            
                            stus = dm.class_students(self.app.data, last_cls)
                            if not any(s["zh"] == zh_name for s in stus):
                                dm.add_student(self.app.data, last_cls, zh_name, en_name, "")
                            
                            try: m_num = int(last_month)
                            except: m_num = month_dict.get(last_month, int(self._mo_var.get()))
                            
                            ym = f"{self._yr_var.get()}-{m_num:02d}"
                            
                            r_date = rows[r]
                            r_range = rows[r+1]
                            r_item = rows[r+2]
                            r_score = rows[r+3]
                            r_retake = rows[r+4]
                            
                            for c in range(6, 27):
                                if c >= len(r_score): break
                                sc = str(r_score[c]).strip() if r_score[c] is not None else ""
                                if sc == "" or sc == "None": continue 
                                
                                day_raw = r_date[c] if c < len(r_date) else ""
                                if day_raw is None or str(day_raw).strip() in ("", "None"): continue
                                
                                try:
                                    day_num = int(float(str(day_raw).strip()))
                                    d_str = f"{self._yr_var.get()}/{m_num:02d}/{day_num:02d}"
                                except:
                                    d_str = str(day_raw).strip()
                                    
                                rng = str(r_range[c]).strip() if c < len(r_range) and r_range[c] is not None else ""
                                itm = str(r_item[c]).strip() if c < len(r_item) and r_item[c] is not None else ""
                                rt = str(r_retake[c]).strip() if c < len(r_retake) and r_retake[c] is not None else ""
                                
                                if rng == "None": rng = ""
                                if itm == "None": itm = ""
                                if rt == "None": rt = ""
                                
                                # 💡 修正 3：移除強制清空 VOC 跟 說 的規則，讓單字跟口說項目能正常寫入
                                if rng == "U": rng = ""
                                if itm == "0": itm = ""
                                
                                if 6 <= c <= 12: t_type = "考試本"
                                elif 13 <= c <= 19: t_type = "口說"
                                elif 20 <= c <= 26: t_type = "單字"
                                else: continue
                                
                                std_val = 90
                                stu_data = next((s for s in stus if s["zh"] == zh_name), None)
                                if stu_data: std_val = stu_data.get("std", 90)
                                
                                rec = {
                                    "type": t_type,
                                    "date": d_str,
                                    "range": rng,
                                    "item": itm,
                                    "score": _clean_val(sc),
                                    "retake": _clean_val(rt),
                                    "std": _clean_val(std_val)
                                }
                                
                                recs = dm.monthly_records(self.app.data, last_cls, zh_name, ym)
                                idx = next((i for i, rx in enumerate(recs) if rx.get("type") == t_type and rx.get("date") == d_str and rx.get("range") == rng and rx.get("item") == itm), None)
                                
                                if idx is not None:
                                    self.app.data.setdefault("monthly", {}).setdefault(last_cls, {}).setdefault(zh_name, {}).setdefault(ym, [])[idx] = rec
                                else:
                                    self.app.data.setdefault("monthly", {}).setdefault(last_cls, {}).setdefault(zh_name, {}).setdefault(ym, []).append(rec)
                                    
                                success_count += 1
                                
                            r += 4 
                    r += 1
            
            if success_count > 0:
                self.app.save()
                # 💡 強制刷新左側班級列表，確保新匯入的班級能選得到
                self.app.page_map["classes"].refresh()
                self.refresh()
                messagebox.showinfo("匯入完成", f"成功從 Excel 同步了 {success_count} 筆成績！\n（若有新班級或新學生，已自動建立完成）", parent=self)
            else:
                messagebox.showwarning("沒有資料", "沒有找到任何成績紀錄，請確認 Excel 格式是否為標準的「兒美成績輸入表」。", parent=self)
                
        except Exception as e:
            messagebox.showerror("匯入失敗", f"讀取 Excel 時發生錯誤：\n{e}", parent=self)

    def _update_html_form(self):
        import sys, re, json
        from pathlib import Path
        base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
        html_names = ["index.html", "老師評語表單.html", "teacher_form.html"]
        html_path = next((str(base / n) for n in html_names if (base / n).exists()), None)

        if not html_path:
            messagebox.showwarning("找不到檔案", "找不到「老師評語表單.html」，請確認檔案與主程式放在同一個資料夾中。", parent=self)
            return

        try:
            t_dict = {}
            s_dict = {}
            for cls, cdata in self.app.data.get("classes", {}).items():
                t = cdata.get("teacher", "未指定")
                b = cdata.get("book", "")
                t_dict.setdefault(t, []).append([cls, b])
                s_dict[cls] = [[s.get("zh", ""), s.get("en", "")] for s in cdata.get("students", [])]

            with open(html_path, "r", encoding="utf-8") as f: 
                html_content = f.read()

            t_str = json.dumps(t_dict, ensure_ascii=False)
            s_str = json.dumps(s_dict, ensure_ascii=False)

            html_content = re.sub(r'const TEACHER_CLASSES\s*=\s*\{.*?\};', f'const TEACHER_CLASSES={t_str};', html_content, flags=re.DOTALL)
            html_content = re.sub(r'const STUDENTS\s*=\s*\{.*?\};', f'const STUDENTS={s_str};', html_content, flags=re.DOTALL)

            with open(html_path, "w", encoding="utf-8") as f: 
                f.write(html_content)
                
            # 產生帶月份參數的連結
            yr = self._yr_var.get()
            mo = self._mo_var.get()
            month_str = yr + "-" + mo
            # 固定使用 GitHub Pages 網址
            # 把老師對應班級清單帶進 URL（JSON base64 編碼）
            import json, base64
            # 老師對應班級：{老師: [[班級, 套書], ...]}
            teacher_map = {}
            for c, cdata in self.app.data.get("classes", {}).items():
                t = cdata.get("teacher", "")
                b = cdata.get("book", "")
                if t:
                    teacher_map.setdefault(t, []).append([c, b])
            teachers_encoded = base64.urlsafe_b64encode(
                json.dumps(teacher_map, ensure_ascii=False).encode()
            ).decode().rstrip("=")
            file_url = ("https://teachergomy-web.github.io/gomyform/?month=" + month_str
                        + "&teachers=" + teachers_encoded)

            # 複製：連結 + LINE 通知文字
            mo_zh = str(int(mo))
            msg_to_copy = (
                "" + mo_zh + " 月評語填寫連結：\n"
                + file_url + "\n\n"
                "請老師們盡快完成補考跟回傳評語🙏"
            )
            self.clipboard_clear()
            self.clipboard_append(msg_to_copy)

            msg = (
                "✅ 已將最新名單同步寫入「老師評語表單.html」\n\n"
                "評語連結（含月份）已複製到剪貼簿，\n"
                "可直接貼到 LINE 傳給老師。\n\n"                
            )
            messagebox.showinfo("更新完成", msg, parent=self)
        except Exception as e:
            messagebox.showerror("更新失敗", f"同步名單至 HTML 時發生錯誤：\n{e}", parent=self)

    def refresh(self):
        cls_list = dm.class_list(self.app.data)
        self._cls_combo["values"] = cls_list
        if cls_list and not self._cls_var.get(): 
            self._cls_combo.current(0)
        self._refresh_table()

    def _quick_edit_header(self, col_key, sample_rec):
        if not sample_rec: 
            return
        dlg = tk.Toplevel(self)
        dlg.title("編輯考試資訊")
        dlg.grab_set()
        dlg.geometry("300x250")
        dlg.update_idletasks()

        f = tk.Frame(dlg, padx=20, pady=20)
        f.pack(fill="both", expand=True)
        U.lbl(f, f"修改「{col_key[0]}」標題資訊", size=16, bold=True).pack(pady=(0, 10))
        
        d_var = tk.StringVar(value=sample_rec.get("date", ""))
        r_var = tk.StringVar(value=sample_rec.get("range", ""))
        i_var = tk.StringVar(value=sample_rec.get("item", ""))
        
        for l, v in [("日期：", d_var), ("範圍：", r_var), ("項目：", i_var)]:
            r = tk.Frame(f)
            r.pack(fill="x", pady=5)
            tk.Label(r, text=l, font=U.FM, width=6).pack(side="left")
            tk.Entry(r, textvariable=v, font=U.FM, width=12).pack(side="left")
        
        def save():
            new_d = d_var.get().strip()
            new_r = r_var.get().strip()
            new_i = i_var.get().strip()
            if not new_d: 
                messagebox.showwarning("錯誤", "日期不能為空", parent=dlg)
                return
            
            ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
            cls = self._cls_var.get()
            updated = False
            
            # 💡 預先記住舊的條件
            old_t = sample_rec.get("type")
            old_d = sample_rec.get("date")
            old_r = sample_rec.get("range")
            old_i = sample_rec.get("item")
            
            # 💡 安全的原始字典更新寫法
            for zh, recs in self.app.data.get("monthly", {}).get(cls, {}).items():
                for rec in recs.get(ym, []):
                    if rec.get("type") == old_t and rec.get("date") == old_d and rec.get("range") == old_r and rec.get("item") == old_i:
                        rec["date"] = new_d
                        rec["range"] = new_r
                        rec["item"] = new_i
                        updated = True
            
            if updated: 
                self.app.save()
            dlg.destroy()
            self._refresh_table()

        U.btn(dlg, "儲存修改", save, size=14).pack(pady=10)

    def _refresh_table(self):
        cls = self._cls_var.get()
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        book = dm.class_book(self.app.data, cls)
        
        if book: 
            self._book_lbl.config(text=f"📚 {book}")
            self._book_lbl.pack(side="left", padx=10)
        else: 
            self._book_lbl.pack_forget()

        for w in self._table_frame.winfo_children(): 
            w.destroy()
            
        if not cls: 
            return

        students = dm.class_students(self.app.data, cls)
        all_recs = dm.all_monthly_records(self.app.data, cls, ym)
        cols = []
        seen = set()
        type_order = ["考試本", "口說", "單字"]
        col_recs_by_zh = {}  
        
        for zh in dm.student_zh_list(self.app.data, cls):
            col_recs_by_zh[zh] = {}
            for rec in all_recs.get(zh, []):
                t = rec.get("type","")
                d = rec.get("date","")
                ri = rec.get("range","")
                it = rec.get("item","")
                key = (t, d, f"{ri} {it}".strip())
                if key not in seen: 
                    seen.add(key)
                    cols.append(key)
                col_recs_by_zh[zh][key] = rec
                
        cols.sort(key=lambda k: (type_order.index(k[0]) if k[0] in type_order else 99, k[1]))
        
        if not cols:
            U.lbl(self._table_frame, "本月尚無成績資料", size=15, fg=U.C["text_lt"], bg=U.C["white"]).grid(row=0, column=0, padx=20, pady=20)
            self._missing_lbl.config(text="")
            return

        self._table_frame.columnconfigure(0, minsize=55)
        for i in range(len(cols)): 
            self._table_frame.columnconfigure(i + 1, minsize=60)

        from itertools import groupby
        type_groups = [(t, len(list(grp))) for t, grp in groupby(cols, key=lambda k: k[0])]

        def th(text, r, c, cs=1, rs=1):
            f = tk.Frame(self._table_frame, bg=U.C["header_bg"], highlightbackground=U.C["border"], highlightthickness=1)
            f.grid(row=r, column=c, columnspan=cs, rowspan=rs, sticky="nsew")
            tk.Label(f, text=text, font=U.FMB, bg=U.C["header_bg"], fg=U.C["primary"], padx=6, pady=4).pack(expand=True)
            return f

        th("學生", 0, 0, rs=2)
        col_offset = 1
        for t, cnt in type_groups: 
            th(t, 0, col_offset, cs=cnt)
            col_offset += cnt
        
        for ci, col_key in enumerate(cols):
            t, d, label = col_key
            f = tk.Frame(self._table_frame, bg=U.C["header_bg"], highlightbackground=U.C["border"], highlightthickness=1)
            f.grid(row=1, column=ci + 1, sticky="nsew")
            lbl_text = d[5:] + (f"\n{label}" if label else "")
            lbl = tk.Label(f, text=lbl_text, font=U.FMB, bg=U.C["header_bg"], fg=U.C["primary"], padx=6, pady=4, cursor="hand2")
            lbl.pack(expand=True, fill="both")
            
            sample_rec = next((recs[col_key] for recs in col_recs_by_zh.values() if col_key in recs), None)
            lbl.bind("<Button-1>", lambda e, k=col_key, r=sample_rec: self._quick_edit_header(k, r))

        missing_list = []
        absent_list = []  
        
        for ri, stu in enumerate(students):
            zh = stu["zh"]
            en = stu["en"]
            row_num = ri + 2
            bg_color = U.C["row_alt"] if ri % 2 == 0 else U.C["white"]
            
            f = tk.Frame(self._table_frame, bg=bg_color, highlightbackground=U.C["border"], highlightthickness=1)
            f.grid(row=row_num, column=0, sticky="nsew")
            tk.Label(f, text=zh, font=U.FMB, bg=bg_color, fg=U.C["text"], padx=8, pady=4).pack(anchor="w")
            tk.Label(f, text=en, font=U.FS, bg=bg_color, fg=U.C["text_lt"], padx=8).pack(anchor="w")

            for ci, col_key in enumerate(cols):
                rec = col_recs_by_zh[zh].get(col_key)
                f_cell = tk.Frame(self._table_frame, highlightbackground=U.C["border"], highlightthickness=1)
                f_cell.grid(row=row_num, column=ci + 1, sticky="nsew")
                
                is_speaking = (col_key[0] == "口說")
                
                if not rec or str(rec.get("score", "")) == "":
                    absent_list.append(f"{zh}({col_key[0]})")
                    lbl = tk.Label(f_cell, text="缺考", font=U.FMB, bg=U.C["white"], fg=U.C["ng_fg"], cursor="hand2")
                    lbl.pack(expand=True, fill="both")
                    lbl.bind("<Button-1>", lambda e, z=zh, k=col_key: self._quick_edit(z, k))
                    continue

                sc = _fmt(rec.get("score", ""))
                rt = _fmt(rec.get("retake", ""))
                
                std = _fmt(stu.get("std", 90))
                
                def is_ok(v):
                    if is_speaking: 
                        return True
                    try: 
                        return float(v) >= float(std)
                    except: 
                        return v in ("A++", "A+", "A")
                        
                ok = is_ok(sc) if rt == "" else is_ok(rt)
                c_bg = U.C["ok_bg"] if ok else U.C["ng_bg"]
                c_fg = U.C["ok_fg"] if ok else U.C["ng_fg"]
                
                if not ok and not is_speaking: 
                    missing_list.append(f"{zh}({col_key[0]})")

                f_cell.config(bg=c_bg)
                lbl = tk.Label(f_cell, text=str(sc) if rt == "" else f"{sc}/{rt}", font=U.FMB, bg=c_bg, fg=c_fg, cursor="hand2")
                lbl.pack(expand=True, fill="both")
                lbl.bind("<Button-1>", lambda e, z=zh, k=col_key: self._quick_edit(z, k))

        msg = []
        if missing_list: 
            msg.append("需補考：" + "、".join(missing_list))
        if absent_list: 
            msg.append("缺考：" + "、".join(absent_list))
            
        if msg:
            self._missing_lbl.config(text="　|　".join(msg))
        else:
            self._missing_lbl.config(text="本月成績均已達標 ✓")

    def _quick_edit(self, zh, col_key):
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        cls = self._cls_var.get()
        recs = dm.monthly_records(self.app.data, cls, zh, ym)
        idx = next((i for i, r in enumerate(recs) if r.get("type") == col_key[0] and r.get("date") == col_key[1]), None)
        
        if idx is not None:
            rec = recs[idx]
        else:
            rec = {"type": col_key[0], "date": col_key[1], "range": "", "item": col_key[2], "score": "", "retake": "", "std": 90}
        
        dlg = tk.Toplevel(self)
        dlg.title(f"快速編輯：{zh}")
        dlg.grab_set()
        dlg.geometry("320x300")
        dlg.update_idletasks()
        dlg.lift()
        dlg.focus_force()
        
        f = tk.Frame(dlg, padx=20, pady=20)
        f.pack()
        U.lbl(f, f"{col_key[0]} ({col_key[1]})", size=16, bold=True).pack(pady=5)
        
        sv = tk.StringVar(value=str(_fmt(rec.get("score", ""))))
        rv = tk.StringVar(value=str(_fmt(rec.get("retake", ""))))
        
        def_std = self.app.data.get("settings", {}).get("default_std", 90)
        stus = dm.class_students(self.app.data, cls)
        stu_data = next((s for s in stus if s["zh"] == zh), None)
        if stu_data:
            def_std = stu_data.get("std", 90)
            
        stdv = tk.StringVar(value=str(_fmt(def_std)))
        is_speaking = (col_key[0] == "口說")
        
        r1 = tk.Frame(f)
        r1.pack(fill="x", pady=5)
        tk.Label(r1, text="成績：", font=U.FM, width=8).pack(side="left")
        tk.Entry(r1, textvariable=sv, font=U.FM, width=10, justify="center").pack(side="left")
        
        if not is_speaking:
            # 補考欄位 (維持可編輯)
            r_rt = tk.Frame(f)
            r_rt.pack(fill="x", pady=5)
            tk.Label(r_rt, text="補考：", font=U.FM, width=8).pack(side="left")
            tk.Entry(r_rt, textvariable=rv, font=U.FM, width=10, justify="center").pack(side="left")
            
            # 標準分欄位 (加入 state="readonly" 設為唯讀)
            r_std = tk.Frame(f)
            r_std.pack(fill="x", pady=5)
            tk.Label(r_std, text="標準分：", font=U.FM, width=8).pack(side="left")
            tk.Entry(r_std, textvariable=stdv, font=U.FM, width=10, justify="center", state="readonly").pack(side="left")

        def save():
            try:
                rec["score"] = _clean_val(sv.get().strip())
                if not is_speaking:
                    rec["retake"] = _clean_val(rv.get().strip())
                    rec["std"] = _clean_val(stdv.get().strip())
                    
                if idx is not None:
                    self.app.data.setdefault("monthly", {}).setdefault(cls, {}).setdefault(zh, {}).setdefault(ym, [])[idx] = rec
                else:
                    self.app.data.setdefault("monthly", {}).setdefault(cls, {}).setdefault(zh, {}).setdefault(ym, []).append(rec)
                
                self.app.save()
            finally:
                dlg.destroy()
                self._refresh_table()

        def delete_record():
            if messagebox.askyesno("確認刪除", "確定要刪除這筆成績紀錄嗎？", parent=dlg):
                if idx is not None:
                    _delete_monthly_rec(self.app.data, cls, zh, ym, rec["type"], rec["date"], rec["range"], rec["item"])
                    self.app.save()
                dlg.destroy()
                self._refresh_table()
                
        br = tk.Frame(dlg)
        br.pack(pady=10)
        U.btn(br, "儲存", save, size=14).pack(side="left", padx=5)
        if idx is not None:
            U.danger_btn(br, "刪除", delete_record, size=14).pack(side="left", padx=5)

# ══════════════════════════════════════════════════════════════════════════════
# 班級 / 學生管理 (支援拖曳排序)
# ══════════════════════════════════════════════════════════════════════════════
class ClassesPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._sel_cls = None
        self._row_widgets = []
        self._build()

    def _build(self):
        U.lbl(self, "班級／學生管理", size=18, bold=True).pack(anchor="w", padx=20, pady=(14, 10))
        body = tk.Frame(self, bg=U.C["bg"])
        body.pack(fill="both", expand=True, padx=20, pady=(0, 12))
        
        left = tk.Frame(body, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1, width=260)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)
        
        lhdr = tk.Frame(left, bg=U.C["header_bg"])
        lhdr.pack(fill="x")
        U.lbl(lhdr, "班級列表", size=14, bold=True, bg=U.C["header_bg"]).pack(side="left", padx=10, pady=7)
        
        # 💡 使用 Treeview 取代 Listbox，並設定 rowheight 來撐開行高
        style = ttk.Style()
        style.configure("ClassList.Treeview", font=U.FM, rowheight=35, borderwidth=0) # 👈 行高在這裡！
        style.layout("ClassList.Treeview", [('ClassList.Treeview.treearea', {'sticky': 'nswe'})]) # 👈 徹底隱藏原生黑框
        
        self._cls_lb = ttk.Treeview(left, style="ClassList.Treeview", show="tree", selectmode="browse")
        self._cls_lb.pack(fill="both", expand=True, padx=4, pady=4)
        self._cls_lb.bind("<<TreeviewSelect>>", self._on_cls_select)
        
        lftr = tk.Frame(left, bg=U.C["white"])
        lftr.pack(fill="x", padx=6, pady=6)
        U.btn(lftr, "新增班級", self._add_cls, size=13).pack(side="left", padx=2)
        U.ghost_btn(lftr, "改名", self._rename_cls, size=13).pack(side="left", padx=2)
        U.danger_btn(lftr, "刪除", self._del_cls, size=13).pack(side="left", padx=2)

        self._right = tk.Frame(body, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        self._right.pack(side="left", fill="both", expand=True)
        
        self._rhdr = tk.Frame(self._right, bg=U.C["header_bg"])
        self._rhdr.pack(fill="x")
        self._rhdr_lbl = U.lbl(self._rhdr, "← 請選擇班級", size=14, bold=True, bg=U.C["header_bg"])
        self._rhdr_lbl.pack(side="left", padx=10, pady=7)
        
        self._add_stu_btn = U.btn(self._rhdr, "新增學生", self._add_stu, size=12)
        self._book_combo_var = tk.StringVar()
        self._book_combo = U.combo(self._rhdr, self._book_combo_var, [], width=10)
        self._book_combo.bind("<<ComboboxSelected>>", self._on_book_change)
        # 授課老師下拉選單
        self._teacher_var = tk.StringVar()
        self._teacher_combo = U.combo(self._rhdr, self._teacher_var, [], width=8)
        self._teacher_combo.bind("<<ComboboxSelected>>", self._on_teacher_change)
        self._teacher_lbl = U.lbl(self._rhdr, "老師：", size=12, bg=U.C["header_bg"])
        
        self._stu_header = tk.Frame(self._right, bg=U.C["header_bg"])
        self._stu_header.pack(fill="x")
        
        self._stu_wrap = tk.Frame(self._right, bg=U.C["white"])
        self._stu_wrap.pack(fill="both", expand=True)
        self._stu_canvas = tk.Canvas(self._stu_wrap, bg=U.C["white"], highlightthickness=0)
        ysb = ttk.Scrollbar(self._stu_wrap, orient="vertical", command=self._stu_canvas.yview)
        self._stu_canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        self._stu_canvas.pack(side="left", fill="both", expand=True)
        self._stu_inner = tk.Frame(self._stu_canvas, bg=U.C["white"])
        self._stu_canvas.create_window((0, 0), window=self._stu_inner, anchor="nw")
        self._stu_inner.bind("<Configure>", lambda e: self._stu_canvas.configure(scrollregion=self._stu_canvas.bbox("all")))
        U.bind_mousewheel(self._stu_canvas)

    def refresh(self): 
        self._load_cls_list()
        self._update_book_combo()
        self._update_teacher_combo()

    def _update_teacher_combo(self):
        if hasattr(self, '_teacher_combo'):
            self._teacher_combo["values"] = dm.teacher_list(self.app.data)

    def _update_book_combo(self): 
        self._book_combo["values"] = dm.book_names(self.app.data)

    def _load_cls_list(self):
        self._cls_lb.delete(*self._cls_lb.get_children())
        for cls in dm.class_list(self.app.data): 
            teacher = dm.class_teacher(self.app.data, cls)
            teacher_str = f"　{teacher}" if teacher else ""
            text_info = f"  {cls}　{dm.class_book(self.app.data, cls)}　{len(dm.class_students(self.app.data, cls))}人{teacher_str}"
            self._cls_lb.insert("", "end", iid=cls, text=text_info)

    def _sel_cls_name(self):
        selected = self._cls_lb.selection()
        return selected[0] if selected else None

    def _on_cls_select(self, event=None):
        cls = self._sel_cls_name()
        if not cls: 
            return
        self._sel_cls = cls
        self._rhdr_lbl.config(text=f"{cls} 學生名單")
        self._book_combo_var.set(dm.class_book(self.app.data, cls))
        self._add_stu_btn.pack(side="right", padx=6, pady=4)
        self._teacher_combo["values"] = dm.teacher_list(self.app.data)
        self._teacher_combo.pack(side="right", padx=2, pady=4)
        self._teacher_lbl.pack(side="right", padx=(6,0), pady=4)
        self._book_combo.pack(side="right", padx=2, pady=4)
        self._teacher_var.set(dm.class_teacher(self.app.data, cls))
        self._load_stu_list()

    def _on_teacher_change(self, event=None):
        if not self._sel_cls:
            return
        self.app.data["classes"][self._sel_cls]["teacher"] = self._teacher_var.get().strip()
        self.app.save()
        self._load_cls_list()

    def _on_book_change(self, event=None):
        if not self._sel_cls: 
            return
        self.app.data["classes"][self._sel_cls]["book"] = self._book_combo_var.get()
        self.app.save()
        self._load_cls_list()

    def _load_stu_list(self):
        for w in self._stu_header.winfo_children(): 
            w.destroy()
        for w in self._stu_inner.winfo_children(): 
            w.destroy()
            
        if not self._sel_cls: 
            return
            
        self._row_widgets = []
        col_defs = [("學生", 0, 160), ("預設標準", 1, 100), ("操作", 2, 240), ("排序", 3, 100)]
        
        for t, ci, mw in col_defs:
            self._stu_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._stu_header, text=t, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=12)

        for i, stu in enumerate(dm.class_students(self.app.data, self._sel_cls)):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row = tk.Frame(self._stu_inner, bg=bg)
            row.pack(fill="x")
            U.sep(self._stu_inner, color=U.C["border"]).pack(fill="x")
            self._row_widgets.append(row)
            
            for _, ci, mw in col_defs: 
                row.columnconfigure(ci, minsize=mw)

            nc = tk.Frame(row, bg=bg)
            nc.grid(row=0, column=0, sticky="w", padx=12, pady=6)
            tk.Label(nc, text=stu["zh"], font=U.FMB, bg=bg, fg=U.C["text"], anchor="w", width=10).pack(anchor="w")
            tk.Label(nc, text=stu["en"], font=U.FS, bg=bg, fg=U.C["text_lt"], anchor="w", width=10).pack(anchor="w")
            
            tk.Label(row, text=stu.get("std", 90), font=U.FMB, bg=bg, fg=U.C["primary"]).grid(row=0, column=1, sticky="w", padx=12)

            oc = tk.Frame(row, bg=bg)
            oc.grid(row=0, column=2, sticky="w", padx=12, pady=6)
            zh = stu["zh"]
            U.ghost_btn(oc, "編輯", lambda z=zh: self._edit_stu(z), size=12).pack(side="left", padx=3)
            U.ghost_btn(oc, "🔀 換班", lambda z=zh: self._transfer_stu(z), size=12).pack(side="left", padx=3)
            U.danger_btn(oc, "離班", lambda z=zh: self._remove_stu(z), size=12).pack(side="left", padx=3)

            sc = tk.Frame(row, bg=bg)
            sc.grid(row=0, column=3, sticky="w", padx=12, pady=6)
            U.ghost_btn(sc, "▲", lambda idx=i: self._move_stu(idx, idx-1), size=11).pack(side="left", padx=1)
            U.ghost_btn(sc, "▼", lambda idx=i: self._move_stu(idx, idx+1), size=11).pack(side="left", padx=1)
            
            handle = tk.Label(sc, text="☰", font=(U.F, 16), bg=bg, fg=U.C["text_lt"], cursor="sb_v_double_arrow")
            handle.pack(side="left", padx=8)
            handle.bind("<Button-1>", lambda e, idx=i: self._on_drag_start(e, idx))
            handle.bind("<ButtonRelease-1>", self._on_drag_release)

    def _move_stu(self, from_idx, to_idx):
        students = self.app.data["classes"][self._sel_cls]["students"]
        if to_idx < 0 or to_idx >= len(students) or from_idx == to_idx: 
            return
        stu = students.pop(from_idx)
        students.insert(to_idx, stu)
        self.app.save()
        self._load_stu_list()

    def _on_drag_start(self, event, idx): 
        self._drag_idx = idx

    def _on_drag_release(self, event):
        if getattr(self, "_drag_idx", None) is None: 
            return
        min_dist = float('inf')
        target_idx = self._drag_idx
        for i, rw in enumerate(self._row_widgets):
            y_center = rw.winfo_rooty() + rw.winfo_height() / 2
            dist = abs(event.y_root - y_center)
            if dist < min_dist: 
                min_dist = dist
                target_idx = i
        self._move_stu(self._drag_idx, target_idx)
        self._drag_idx = None

    def _add_cls(self):
        name = U.ask_string(self, "新增班級", "班級名稱（例如：A班）：")
        if not name: 
            return
        if name in self.app.data["classes"]: 
            messagebox.showwarning("重複", f"「{name}」已存在", parent=self)
            return
        dm.add_class(self.app.data, name)
        self.app.save()
        self._load_cls_list()

    def _rename_cls(self):
        cls = self._sel_cls_name()
        if not cls: 
            return
        new = U.ask_string(self, "班級改名", "新名稱：", default=cls)
        if not new or new == cls: 
            return
        if new in self.app.data["classes"]: 
            messagebox.showwarning("重複", f"「{new}」已存在", parent=self)
            return
        dm.rename_class(self.app.data, cls, new)
        self.app.save()
        self._sel_cls = new
        self._load_cls_list()

    def _del_cls(self):
        cls = self._sel_cls_name()
        if not cls: 
            return
        if not U.ask_yesno(self, "確認刪除", f"確定刪除「{cls}」？\n（學生名單會移除，歷史成績保留）"): 
            return
        dm.delete_class(self.app.data, cls)
        self.app.save()
        self._sel_cls = None
        self._load_cls_list()
        for w in self._stu_inner.winfo_children(): 
            w.destroy()

    def _stu_dialog(self, title, zh="", en="", std=None):
        if std is None:
            std = self.app.data.get("settings", {}).get("default_std", 90)
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.after(50, lambda: (dlg.update_idletasks(), dlg.lift(), dlg.focus_force()))
        
        res = [None]
        f = tk.Frame(dlg, padx=20, pady=16)
        f.pack()
        
        zh_v = tk.StringVar(value=zh)
        en_v = tk.StringVar(value=en)
        std_v = tk.StringVar(value=str(std))
        
        for l, v in [("中文姓名：", zh_v), ("英文姓名：", en_v)]:
            r = tk.Frame(f)
            r.pack(fill="x", pady=3)
            tk.Label(r, text=l, font=U.FM, width=10, anchor="w").pack(side="left")
            tk.Entry(r, textvariable=v, font=U.FM, width=18).pack(side="left")
        
        r = tk.Frame(f)
        r.pack(fill="x", pady=3)
        tk.Label(r, text="預設標準分：", font=U.FM, width=10, anchor="w").pack(side="left")
        tk.Entry(r, textvariable=std_v, font=U.FM, width=10).pack(side="left")

        def ok():
            try: 
                s = int(std_v.get())
            except Exception: 
                s = 90
            res[0] = (zh_v.get().strip(), en_v.get().strip(), "", s)
            dlg.destroy()
            
        br = tk.Frame(dlg)
        br.pack(pady=8)
        U.btn(br, "確定", ok, size=13, width=10).pack(side="left", padx=5)
        U.ghost_btn(br, "取消", dlg.destroy, size=13).pack(side="left", padx=5)
        
        dlg.update_idletasks()
        dlg.lift()
        dlg.focus_force()
        self.wait_window(dlg)
        return res[0]

    def _add_stu(self):
        if not self._sel_cls: 
            return
        res = self._stu_dialog("新增學生")
        if not res: 
            return
        zh, en, _, std = res
        if any(s["zh"] == zh for s in dm.class_students(self.app.data, self._sel_cls)): 
            messagebox.showwarning("重複", "名單中已存在", parent=self)
            return
            
        self.app.data["classes"][self._sel_cls]["students"].append({"zh": zh, "en": en, "std": std})
        self.app.save()
        self._load_stu_list()

    def _edit_stu(self, zh):
        students = dm.class_students(self.app.data, self._sel_cls)
        stu = next((s for s in students if s["zh"] == zh), None)
        if not stu: 
            return
            
        res = self._stu_dialog("編輯學生", stu["zh"], stu.get("en", ""), stu.get("std", 90))
        if not res: 
            return
            
        new_zh, new_en, _, new_std = res
        stu["zh"] = new_zh
        stu["en"] = new_en
        stu["std"] = new_std
        self.app.save()
        self._load_stu_list()

    def _transfer_stu(self, zh):
        others = [c for c in dm.class_list(self.app.data) if c != self._sel_cls]
        target = U.ask_choice(self, "換班", f"移到哪個班？", others)
        if target: 
            dm.transfer_student(self.app.data, self._sel_cls, target, zh)
            self.app.save()
            self._load_stu_list()

    def _remove_stu(self, zh):
        if U.ask_yesno(self, "確認離班", f"確定讓「{zh}」離班？"): 
            dm.remove_student(self.app.data, self._sel_cls, zh)
            self.app.save()
            self._load_stu_list()

# ══════════════════════════════════════════════════════════════════════════════
# 月成績輸入 (詳細模式) 💡隱藏年月、顯示用書、修正儲存、支援自動滾輪
# ══════════════════════════════════════════════════════════════════════════════
class MonthlyInputPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._cls_var = tk.StringVar()
        self._type_var = tk.StringVar(value="考試本")
        self._date_var = tk.StringVar(value=datetime.now().strftime("%Y/%m/%d"))
        self._range_var = tk.StringVar()
        self._item_var = tk.StringVar()
        self._rows = []
        self._row_widgets = []
        self._build()

    def _build(self):
        self._topbar = tk.Frame(self, bg=U.C["bg"])
        self._topbar.pack(fill="x", padx=20, pady=(14, 8))
        U.lbl(self._topbar, "月成績詳細輸入模式", size=18, bold=True).pack(side="left")
        U.ghost_btn(self._topbar, "＜ 返回總覽", lambda: self.app._show("overview"), size=13).pack(side="right")

        frow = tk.Frame(self, bg=U.C["bg"])
        frow.pack(fill="x", padx=20, pady=(0, 8))
        
        U.lbl(frow, "班級", size=13, bg=U.C["bg"]).pack(side="left")
        self._cls_combo = U.combo(frow, self._cls_var, [], width=8)
        self._cls_combo.pack(side="left", padx=(4, 12))
        self._cls_combo.bind("<<ComboboxSelected>>", lambda e: self._on_cls_change())
        
        U.lbl(frow, "用書：", size=13, bg=U.C["bg"]).pack(side="left", padx=(10, 2))
        self._book_lbl = U.lbl(frow, "", size=13, bg=U.C["bg"], fg=U.C["primary"], bold=True)
        self._book_lbl.pack(side="left", padx=(0, 12))
        
        tab_frame = tk.Frame(self, bg=U.C["bg"])
        tab_frame.pack(fill="x", padx=20, pady=(0, 6))
        self._tab_btns = {}
        for t in ["考試本", "口說", "單字"]:
            b = tk.Button(tab_frame, text=t, font=U.FM, relief="flat", cursor="hand2", padx=14, pady=6, command=lambda t=t: self._switch_type(t))
            b.pack(side="left")
            self._tab_btns[t] = b
            
        self._update_tabs()

        meta = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        meta.pack(fill="x", padx=20, pady=(0, 8))
        mr = tk.Frame(meta, bg=U.C["white"])
        mr.pack(fill="x", padx=10, pady=8)
        U.lbl(mr, "日期", size=13, bg=U.C["white"]).pack(side="left")
        U.date_btn(mr, self._date_var, bg=U.C["white"]).pack(side="left", padx=(4,12))
        U.lbl(mr, "範圍", size=13, bg=U.C["white"]).pack(side="left")
        tk.Entry(mr, textvariable=self._range_var, font=U.FM, width=12).pack(side="left", padx=(4,12))
        U.lbl(mr, "項目", size=13, bg=U.C["white"]).pack(side="left")
        tk.Entry(mr, textvariable=self._item_var, font=U.FM, width=10).pack(side="left", padx=(4,0))

        tbl_wrap = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        tbl_wrap.pack(fill="both", expand=True, padx=20)
        self._tbl_header = tk.Frame(tbl_wrap, bg=U.C["header_bg"])
        self._tbl_header.pack(fill="x")
        self._tbl_canvas = tk.Canvas(tbl_wrap, bg=U.C["white"], highlightthickness=0)
        ysb = ttk.Scrollbar(tbl_wrap, orient="vertical", command=self._tbl_canvas.yview)
        self._tbl_canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        self._tbl_canvas.pack(side="left", fill="both", expand=True)
        self._tbl_inner = tk.Frame(self._tbl_canvas, bg=U.C["white"])
        self._tbl_canvas.create_window((0,0), window=self._tbl_inner, anchor="nw")
        self._tbl_inner.bind("<Configure>", lambda e: self._tbl_canvas.configure(scrollregion=self._tbl_canvas.bbox("all")))
        U.bind_mousewheel(self._tbl_canvas)

        bot = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        bot.pack(fill="x", padx=20, pady=(0, 10))
        self._summary_lbl = U.lbl(bot, "", size=13, bg=U.C["white"], fg=U.C["text_lt"])
        self._summary_lbl.pack(side="left", padx=12, pady=7)
        U.btn(bot, "儲存全部 ✓", lambda: self._save_all(show_msg=True), size=13).pack(side="right", padx=8, pady=5)

    def _on_cls_change(self):
        cls = self._cls_var.get()
        book = dm.class_book(self.app.data, cls)
        self._book_lbl.config(text=book if book else "無")
        self._render_rows()

    def _switch_type(self, t): 
        self._type_var.set(t)
        self._update_tabs()
        self._render_rows()
        
    def _update_tabs(self):
        t = self._type_var.get()
        for k, b in self._tab_btns.items():
            b.config(bg=U.C["primary"] if k==t else U.C["bg"], fg="white" if k==t else U.C["text_lt"])

    def refresh(self):
        cls_list = dm.class_list(self.app.data)
        self._cls_combo["values"] = cls_list
        ov_cls = self.app.page_map["overview"]._cls_var.get()
        if ov_cls in cls_list: 
            self._cls_var.set(ov_cls)
        elif cls_list and not self._cls_var.get(): 
            self._cls_combo.current(0)
        self._on_cls_change()

    def _render_rows(self):
        for w in self._tbl_header.winfo_children(): 
            w.destroy()
        for w in self._tbl_inner.winfo_children(): 
            w.destroy()
            
        self._rows = []
        self._row_widgets = []
        cls = self._cls_var.get()
        if not cls: 
            return
            
        has_rt = (self._type_var.get() != "口說")
        col_defs = [("學生", 0, 160), ("成績", 1, 100)]
        
        sort_col = 2
        if has_rt: 
            col_defs.extend([("補考", 2, 100), ("標準", 3, 80)])
            sort_col = 4
        else: 
            col_defs.extend([("標準", 2, 80)])
            sort_col = 3
            
        col_defs.append(("排序", sort_col, 100))

        for t, ci, mw in col_defs:
            self._tbl_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._tbl_header, text=t, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=12)

        num_nav_cols = 2 if has_rt else 1
        nav_matrix = [[] for _ in range(num_nav_cols)]

        for i, stu in enumerate(dm.class_students(self.app.data, cls)):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row = tk.Frame(self._tbl_inner, bg=bg)
            row.pack(fill="x")
            U.sep(self._tbl_inner, color=U.C["border"]).pack(fill="x")
            self._row_widgets.append(row)
            
            for _, ci, mw in col_defs: 
                row.columnconfigure(ci, minsize=mw)
            
            sv = tk.StringVar()
            rv = tk.StringVar()
            stdv = tk.StringVar(value=str(_fmt(stu.get("std", 90))))
            self._rows.append({"zh": stu["zh"], "score_var": sv, "retake_var": rv, "std_var": stdv})

            nc = tk.Frame(row, bg=bg)
            nc.grid(row=0, column=0, sticky="w", padx=12, pady=6)
            tk.Label(nc, text=stu["zh"], font=U.FMB, bg=bg, fg=U.C["text"]).pack(anchor="w")
            tk.Label(nc, text=stu["en"], font=U.FS, bg=bg, fg=U.C["text_lt"]).pack(anchor="w")
            
            is_speak_type = (self._type_var.get() == "口說")
            if is_speak_type: 
                sc_w = tk.Entry(row, textvariable=sv, font=U.FM, width=8, justify="center")
            else:
                vcmd = _make_score_validator(self)
                sc_w = tk.Entry(row, textvariable=sv, font=U.FM, width=8, justify="center", validate="key", validatecommand=vcmd)
                
            sc_w.grid(row=0, column=1, padx=12)
            nav_matrix[0].append(sc_w)

            if has_rt:
                if is_speak_type: 
                    rt_w = tk.Entry(row, textvariable=rv, font=U.FM, width=8, justify="center")
                else:
                    vcmd = _make_score_validator(self)
                    rt_w = tk.Entry(row, textvariable=rv, font=U.FM, width=8, justify="center", validate="key", validatecommand=vcmd)
                    
                rt_w.grid(row=0, column=2, padx=12)
                nav_matrix[1].append(rt_w)
                tk.Entry(row, textvariable=stdv, font=U.FM, width=6, justify="center").grid(row=0, column=3, padx=12)
            else:
                tk.Entry(row, textvariable=stdv, font=U.FM, width=6, justify="center").grid(row=0, column=2, padx=12)

            sc = tk.Frame(row, bg=bg)
            sc.grid(row=0, column=sort_col, sticky="w", padx=12, pady=6)
            U.ghost_btn(sc, "▲", lambda idx=i: self._move_stu(idx, idx-1), size=12).pack(side="left", padx=1)
            U.ghost_btn(sc, "▼", lambda idx=i: self._move_stu(idx, idx+1), size=12).pack(side="left", padx=1)
            handle = tk.Label(sc, text="☰", font=(U.F, 16), bg=bg, fg=U.C["text_lt"], cursor="sb_v_double_arrow")
            handle.pack(side="left", padx=8)
            handle.bind("<Button-1>", lambda e, idx=i: self._on_drag_start(e, idx))
            handle.bind("<ButtonRelease-1>", self._on_drag_release)

        # 💡 啟用智慧滾動導航
        apply_matrix_nav(nav_matrix, canvas=self._tbl_canvas)

    def _move_stu(self, from_idx, to_idx):
        cls = self._cls_var.get()
        students = self.app.data["classes"][cls]["students"]
        if to_idx < 0 or to_idx >= len(students) or from_idx == to_idx: 
            return
        self._save_all(show_msg=False) 
        stu = students.pop(from_idx)
        students.insert(to_idx, stu)
        self.app.save()
        self._render_rows()

    def _on_drag_start(self, event, idx): 
        self._drag_idx = idx

    def _on_drag_release(self, event):
        if getattr(self, "_drag_idx", None) is None: 
            return
        min_dist = float('inf')
        target_idx = self._drag_idx
        for i, rw in enumerate(self._row_widgets):
            y_center = rw.winfo_rooty() + rw.winfo_height() / 2
            dist = abs(event.y_root - y_center)
            if dist < min_dist: 
                min_dist = dist
                target_idx = i
        self._move_stu(self._drag_idx, target_idx)
        self._drag_idx = None

    def _save_all(self, show_msg=True):
        cls = self._cls_var.get()
        ov = self.app.page_map["overview"]
        ym = f"{ov._yr_var.get()}-{ov._mo_var.get()}"
        dt = self._date_var.get().strip()
        rg = self._range_var.get().strip()
        it = self._item_var.get().strip()
        
        if not dt or not it:
            messagebox.showwarning("提示", "請填寫日期與項目", parent=self)
            return

        saved_any = False
        for r in self._rows:
            sc = r["score_var"].get().strip()
            
            recs = dm.monthly_records(self.app.data, cls, r["zh"], ym)
            idx = next((i for i, rx in enumerate(recs) if rx.get("type") == self._type_var.get() and rx.get("date") == dt and rx.get("range") == rg and rx.get("item") == it), None)

            # 💡 若清空分數且原本有紀錄，執行安全刪除
            if not sc:
                if idx is not None:
                    _delete_monthly_rec(self.app.data, cls, r["zh"], ym, self._type_var.get(), dt, rg, it)
                    saved_any = True
                continue
            
            rec = {
                "type": self._type_var.get(), 
                "date": dt, 
                "range": rg, 
                "item": it, 
                "score": _clean_val(sc), 
                "retake": _clean_val(r["retake_var"].get().strip()), 
                "std": _clean_val(r["std_var"].get())
            }
            if idx is not None:
                self.app.data.setdefault("monthly", {}).setdefault(cls, {}).setdefault(r["zh"], {}).setdefault(ym, [])[idx] = rec
            else:
                self.app.data.setdefault("monthly", {}).setdefault(cls, {}).setdefault(r["zh"], {}).setdefault(ym, []).append(rec)
            saved_any = True
            
        self.app.save()
        
        if show_msg: 
            if saved_any: 
                messagebox.showinfo("完成", "已成功儲存/刪除變更", parent=self)
            else: 
                messagebox.showinfo("提示", "沒有填寫任何成績", parent=self)
            
        self._render_rows()

# ══════════════════════════════════════════════════════════════════════════════
# 用書管理
# ══════════════════════════════════════════════════════════════════════════════
class BooksPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._build()

    def _build(self):
        U.lbl(self, "用書管理", size=18, bold=True).pack(anchor="w", padx=20, pady=(14, 4))
        U.lbl(self, "套書升級順序由上到下固定，點「編輯」可修改各套考試項目", size=12, fg=U.C["text_lt"]).pack(anchor="w", padx=20, pady=(0, 10))

        self._list_frame = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        self._list_frame.pack(fill="both", expand=True, padx=20)

        self._bks_header = tk.Frame(self._list_frame, bg=U.C["header_bg"])
        self._bks_header.pack(fill="x")

        self._rows_frame = tk.Frame(self._list_frame, bg=U.C["white"])
        self._rows_frame.pack(fill="both", expand=True)

        footer = tk.Frame(self, bg=U.C["bg"])
        footer.pack(fill="x", padx=20, pady=10)
        U.btn(footer, "新增套書", self._add_book, size=13).pack(side="left")

    def refresh(self):
        self._render_rows()

    def _render_rows(self):
        for w in self._bks_header.winfo_children(): 
            w.destroy()
        for w in self._rows_frame.winfo_children(): 
            w.destroy()
        
        col_defs = [("  #", 0, 60), ("套書名稱", 1, 160), ("升級考試項目", 2, 350), ("", 3, 100)]
        for text, ci, mw in col_defs:
            self._bks_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._bks_header, text=text, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=10)

        for i, book in enumerate(self.app.data["books"]):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row = tk.Frame(self._rows_frame, bg=bg)
            row.pack(fill="x")
            U.sep(self._rows_frame, color=U.C["border"]).pack(fill="x")
            
            for _, ci, mw in col_defs: 
                row.columnconfigure(ci, minsize=mw)

            tk.Label(row, text=str(i+1), font=U.FM, bg=bg, fg=U.C["text_lt"]).grid(row=0, column=0, sticky="w", padx=10, pady=6)
            tk.Label(row, text=book["name"], font=U.FMB, bg=bg, fg=U.C["text"]).grid(row=0, column=1, sticky="w", padx=10, pady=6)
            tk.Label(row, text="・".join(book["items"]), font=U.FM, bg=bg, fg=U.C["text_lt"]).grid(row=0, column=2, sticky="w", padx=10, pady=6)
            
            bc = tk.Frame(row, bg=bg)
            bc.grid(row=0, column=3, sticky="w", padx=10, pady=6)
            U.ghost_btn(bc, "編輯", lambda idx=i: self._edit_book(idx), size=13).pack()

    def _add_book(self):
        self._show_book_dialog()

    def _edit_book(self, idx):
        self._show_book_dialog(idx)

    def _show_book_dialog(self, idx=None):
        is_edit = (idx is not None)
        if is_edit:
            book = self.app.data["books"][idx]
            title = f"編輯：{book['name']}"
            b_name = book["name"]
            b_items = book["items"]
        else:
            title = "新增套書"
            b_name = ""
            b_items = [] 

        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.after(50, lambda: (dlg.update_idletasks(), dlg.lift(), dlg.focus_force()))
        
        f = tk.Frame(dlg, padx=20, pady=16)
        f.pack()
        
        name_v = tk.StringVar(value=b_name)
        r = tk.Frame(f)
        r.pack(fill="x", pady=4)
        tk.Label(r, text="套書名稱：", font=U.FM, width=10, anchor="w").pack(side="left")
        tk.Entry(r, textvariable=name_v, font=U.FM, width=16).pack(side="left")

        tk.Label(f, text="考試項目（勾選）：", font=U.FMB, anchor="w").pack(anchor="w", pady=(10,4))
        all_items = ["聽力", "單字", "字母", "翻譯", "口說"]
        item_vars = {}
        for it in all_items:
            v = tk.BooleanVar(value=(it in b_items))
            item_vars[it] = v
            tk.Checkbutton(f, text=it, variable=v, font=U.FM).pack(anchor="w")

        def save():
            new_name = name_v.get().strip()
            if not new_name: 
                messagebox.showwarning("提示", "請填寫名稱", parent=dlg)
                return
                
            for i, b in enumerate(self.app.data["books"]):
                if b["name"] == new_name and i != idx:
                    messagebox.showwarning("重複", "名稱已存在", parent=dlg)
                    return
                    
            items = [it for it in all_items if item_vars[it].get()]
            if not items: 
                messagebox.showwarning("提示", "至少選一個考試項目", parent=dlg)
                return
                
            if is_edit:
                old_name = self.app.data["books"][idx]["name"]
                self.app.data["books"][idx]["name"] = new_name
                self.app.data["books"][idx]["items"] = items
                for c_name, c_data in self.app.data.get("classes", {}).items():
                    if c_data.get("book") == old_name:
                        c_data["book"] = new_name
            else:
                self.app.data["books"].append({"name": new_name, "items": items})
                
            self.app.save()
            dlg.destroy()
            self._render_rows()

        def delete_book():
            book_name = self.app.data["books"][idx]["name"]
            using_classes = [c for c, cdata in self.app.data.get("classes", {}).items() if cdata.get("book") == book_name]
            if using_classes:
                messagebox.showwarning("無法刪除", f"以下班級正在使用此套書，請先修改班級用書：\n{', '.join(using_classes)}", parent=dlg)
                return
                
            if not messagebox.askyesno("確認刪除", f"確定要刪除「{book_name}」套書嗎？", parent=dlg):
                return
                
            del self.app.data["books"][idx]
            self.app.save()
            dlg.destroy()
            self._render_rows()

        br = tk.Frame(dlg)
        br.pack(pady=8)
        U.btn(br, "確定", save, size=13).pack(side="left", padx=5)
        if is_edit:
            U.danger_btn(br, "刪除", delete_book, size=13).pack(side="left", padx=5)
        U.ghost_btn(br, "取消", dlg.destroy, size=13).pack(side="left", padx=5)

# ══════════════════════════════════════════════════════════════════════════════
# 升級考成績輸入 (無缺考警告、支援智慧導航、小數點淨化)
# ══════════════════════════════════════════════════════════════════════════════
class LevelTestPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._in_cls_var  = tk.StringVar()
        self._in_date_var = tk.StringVar(value=datetime.now().strftime("%Y/%m/%d"))
        self._rows = []

        self._build_overview()
        self._build_input()
        self._show_overview()

    def _show_overview(self):
        self._input_frame.pack_forget()
        self._ov_frame.pack(fill="both", expand=True)
        self.refresh()
        self.update_idletasks()

    def _show_input(self, cls="", date=""):
        self._ov_frame.pack_forget()
        self._input_frame.pack(fill="both", expand=True)

        cls_list = dm.class_list(self.app.data)
        self._in_cls_combo["values"] = cls_list

        if cls: 
            self._in_cls_var.set(cls)
        elif cls_list: 
            self._in_cls_var.set(cls_list[0])
        else: 
            self._in_cls_var.set("")

        if date: 
            self._in_date_var.set(date)
        else: 
            self._in_date_var.set(datetime.now().strftime("%Y/%m/%d"))

        self._render_input_rows()
        self.update_idletasks()

    def _build_overview(self):
        self._ov_frame = tk.Frame(self, bg=U.C["bg"])

        topbar = tk.Frame(self._ov_frame, bg=U.C["bg"])
        topbar.pack(fill="x", padx=20, pady=(14, 8))
        U.lbl(topbar, "升級考成績輸入", size=18, bold=True).pack(side="left")
        U.btn(topbar, "新增成績紀錄", lambda: self._show_input(), size=13).pack(side="right", padx=10)

        tbl_wrap = tk.Frame(self._ov_frame, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        tbl_wrap.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        self._ov_header = tk.Frame(tbl_wrap, bg=U.C["header_bg"])
        self._ov_header.pack(fill="x")
        self._ov_canvas = tk.Canvas(tbl_wrap, bg=U.C["white"], highlightthickness=0)
        ysb = ttk.Scrollbar(tbl_wrap, orient="vertical", command=self._ov_canvas.yview)
        self._ov_canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        self._ov_canvas.pack(side="left", fill="both", expand=True)
        self._ov_inner = tk.Frame(self._ov_canvas, bg=U.C["white"])
        self._ov_canvas.create_window((0,0), window=self._ov_inner, anchor="nw")
        self._ov_inner.bind("<Configure>", lambda e: self._ov_canvas.configure(scrollregion=self._ov_canvas.bbox("all")))
        U.bind_mousewheel(self._ov_canvas)

    def _build_input(self):
        self._input_frame = tk.Frame(self, bg=U.C["bg"])

        topbar = tk.Frame(self._input_frame, bg=U.C["bg"])
        topbar.pack(fill="x", padx=20, pady=(14, 8))
        U.lbl(topbar, "升級考詳細輸入模式", size=18, bold=True).pack(side="left")
        U.ghost_btn(topbar, "＜ 返回總覽", lambda: self._show_overview(), size=13).pack(side="right")

        frow = tk.Frame(self._input_frame, bg=U.C["bg"])
        frow.pack(fill="x", padx=20, pady=(0, 8))
        U.lbl(frow, "班級", size=13, bg=U.C["bg"]).pack(side="left")
        self._in_cls_combo = U.combo(frow, self._in_cls_var, [], width=8)
        self._in_cls_combo.pack(side="left", padx=(4, 12))
        self._in_cls_combo.bind("<<ComboboxSelected>>", lambda e: self._render_input_rows())

        U.lbl(frow, "考試日期", size=13, bg=U.C["bg"]).pack(side="left")
        U.date_btn(frow, self._in_date_var).pack(side="left", padx=(4, 12))
        self._in_date_var.trace_add("write", lambda *a: self._render_input_rows())

        self._in_info_lbl = U.lbl(frow, "", size=14, bg=U.C["bg"], fg=U.C["text_lt"])
        self._in_info_lbl.pack(side="left", padx=12)
        U.ghost_btn(frow, "↺ 清空重填", self._clear_inputs, size=13).pack(side="right", padx=4)

        tbl_wrap = tk.Frame(self._input_frame, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        tbl_wrap.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        self._in_header = tk.Frame(tbl_wrap, bg=U.C["header_bg"])
        self._in_header.pack(fill="x")
        self._in_canvas = tk.Canvas(tbl_wrap, bg=U.C["white"], highlightthickness=0)
        ysb = ttk.Scrollbar(tbl_wrap, orient="vertical", command=self._in_canvas.yview)
        self._in_canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        self._in_canvas.pack(side="left", fill="both", expand=True)
        self._in_inner = tk.Frame(self._in_canvas, bg=U.C["white"])
        self._in_canvas.create_window((0,0), window=self._in_inner, anchor="nw")
        self._in_inner.bind("<Configure>", lambda e: self._in_canvas.configure(scrollregion=self._in_canvas.bbox("all")))
        U.bind_mousewheel(self._in_canvas)
        
        self._in_bot = tk.Frame(self._input_frame, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        self._in_bot.pack(fill="x", padx=20, pady=(0, 10))
        
        U.btn(self._in_bot, "儲存全部 ✓", self._save_all, size=13).pack(side="right", padx=8, pady=5)

    def refresh(self):
        self._render_overview_rows()

    def _render_overview_rows(self):
        for w in self._ov_header.winfo_children(): 
            w.destroy()
        for w in self._ov_inner.winfo_children(): 
            w.destroy()

        records = []
        for cls, cdata in self.app.data.get("classes", {}).items():
            lts = cdata.get("leveltests", [])
            if not lts: 
                lts = self.app.data.get("leveltest", {}).get(cls, [])
            for lt in lts:
                records.append({"date": lt.get("date", ""), "cls": cls, "book": lt.get("book", ""), "data": lt})

        records.sort(key=lambda x: (x["date"], x["cls"]), reverse=True)

        col_defs = [("考試日期", 0, 180), ("班級", 1, 120), ("套書", 2, 120), ("操作", 3, 280)]
        for t, ci, mw in col_defs:
            self._ov_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._ov_header, text=t, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=12)

        if not records:
            U.lbl(self._ov_inner, "目前尚無任何升級考紀錄", size=15, fg=U.C["text_lt"], bg=U.C["white"]).grid(row=0, column=0, padx=20, pady=20)
            return

        for i, rec in enumerate(records):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row = tk.Frame(self._ov_inner, bg=bg)
            row.pack(fill="x")
            U.sep(self._ov_inner, color=U.C["border"]).pack(fill="x")
            
            for _, ci, mw in col_defs: 
                row.columnconfigure(ci, minsize=mw)

            tk.Label(row, text=rec["date"], font=U.FMB, bg=bg, fg=U.C["text"]).grid(row=0, column=0, sticky="w", padx=12, pady=10)
            tk.Label(row, text=rec["cls"], font=U.FMB, bg=bg, fg=U.C["primary"]).grid(row=0, column=1, sticky="w", padx=12, pady=10)
            tk.Label(row, text=rec["book"], font=U.FM, bg=bg, fg=U.C["text_lt"]).grid(row=0, column=2, sticky="w", padx=12, pady=10)

            op_frame = tk.Frame(row, bg=bg)
            op_frame.grid(row=0, column=3, sticky="w", padx=12, pady=10)
            U.ghost_btn(op_frame, "檢視 / 編輯成績", lambda c=rec["cls"], d=rec["date"]: self._show_input(c, d), size=13).pack(side="left", padx=4)
            U.ghost_btn(op_frame, "🔗 評語連結", lambda c=rec["cls"], d=rec["date"]: self._copy_lt_link(c, d), size=13).pack(side="left", padx=4)
            U.danger_btn(op_frame, "刪除", lambda c=rec["cls"], d=rec["date"]: self._delete_lt(c, d), size=13).pack(side="left", padx=4)

    def _copy_lt_link(self, cls, date):
        """複製升級考評語填寫連結"""
        import sys, json, base64
        from pathlib import Path

        # 把分數編碼進 URL，表單可以顯示給老師看
        lts = self.app.data["classes"][cls].get("leveltests", [])
        if not lts:
            lts = self.app.data.get("leveltest", {}).get(cls, [])
        lt = next((x for x in lts if x.get("date") == date), None)

        scores_encoded = ""
        if lt:
            # 整理每個學生的分數 {zh: {關卡: 分數}}
            scores_data = {}
            for zh, zh_scores in lt.get("scores", {}).items():
                scores_data[zh] = {}
                for it, sc in zh_scores.items():
                    if sc != "" and sc is not None:
                        scores_data[zh][it] = str(sc)
                    rt = lt.get("retakes", {}).get(zh, {}).get(it, "")
                    if rt != "" and rt is not None:
                        scores_data[zh][it + "_retake"] = str(rt)
            try:
                # 用 URL-safe base64（+ → - ，/ → _，去掉 =）
                raw = base64.urlsafe_b64encode(
                    json.dumps(scores_data, ensure_ascii=False).encode()
                ).decode().rstrip('=')
                scores_encoded = raw
            except Exception:
                scores_encoded = ""

        book = dm.class_book(self.app.data, cls)
        base_url = "https://teachergomy-web.github.io/gomyform/lt.html"
        # 從 date 取月份（如 2026/04/25 → 2026-04）
        month_from_date = date.replace("/", "-")[:7]
        teacher = dm.class_teacher(self.app.data, cls)
        link = (base_url
                + "?month=" + month_from_date
                + "&cls=" + cls
                + "&book=" + book
                + "&date=" + date.replace("/", "-")
                + ("&teacher=" + teacher if teacher else "")
                + ("&scores=" + scores_encoded if scores_encoded else ""))

        self.clipboard_clear()
        self.clipboard_append(link)

        month = date[:7].replace("/", "-") if "/" in date else date[:7]
        msg = ("已複製 " + cls + " 升級考評語填寫連結\n\n"
               + "請貼到 LINE 傳給老師：\n" + link[:80] + "...")
        messagebox.showinfo("連結已複製", msg, parent=self)

    def _delete_lt(self, cls, date):
        if not messagebox.askyesno("確認刪除", f"確定要刪除 {cls} 於 {date} 的升級考紀錄嗎？\n(將無法復原)"): 
            return
            
        lts = self.app.data["classes"][cls].get("leveltests", [])
        self.app.data["classes"][cls]["leveltests"] = [x for x in lts if x.get("date") != date]
        
        old_lts = self.app.data.get("leveltest", {}).get(cls, [])
        self.app.data.setdefault("leveltest", {})[cls] = [x for x in old_lts if x.get("date") != date]
        
        self.app.save()
        self._render_overview_rows()

    def _clear_inputs(self):
        if not messagebox.askyesno("確認", "確定要清空目前畫面上所有尚未儲存的輸入嗎？"): 
            return
        for r in self._rows:
            for vars_ in r["items"].values():
                vars_["score_var"].set("")
                vars_["retake_var"].set("")

    def _render_input_rows(self):
        for w in self._in_inner.winfo_children(): 
            w.destroy()
        for w in self._in_header.winfo_children(): 
            w.destroy()
            
        self._rows = []
        cls = self._in_cls_var.get()
        date = self._in_date_var.get()
        if not cls or not date: 
            return

        book = dm.class_book(self.app.data, cls)
        items = dm.book_items(self.app.data, book)
        self._in_info_lbl.config(text=f"套書：{book}　　考試項目：{'、'.join(items)}")
        
        if not items:
            U.lbl(self._in_inner, "此班尚未設定套書，請先至「用書管理」設定", size=15, fg=U.C["ng_fg"], bg=U.C["white"]).pack(padx=20, pady=20)
            return

        lts = self.app.data["classes"].get(cls, {}).get("leveltests", [])
        lt = next((x for x in lts if x.get("date") == date), {})

        self._in_header.columnconfigure(0, minsize=160)
        tk.Label(self._in_header, text="學生", font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=0, rowspan=2, sticky="w", padx=12)

        nav_matrix = []
        col_idx = 0
        for idx, it in enumerate(items):
            base_col = 1 + idx * 2
            self._in_header.columnconfigure(base_col, minsize=70)
            self._in_header.columnconfigure(base_col+1, minsize=70)
            
            tk.Label(self._in_header, text=it, font=U.FMB, bg=U.C["header_bg"], fg=U.C["primary"], anchor="center", pady=4).grid(row=0, column=base_col, columnspan=2)
            tk.Label(self._in_header, text="成績", font=U.FS, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="center").grid(row=1, column=base_col, pady=(0,6))
            tk.Label(self._in_header, text="補考", font=U.FS, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="center").grid(row=1, column=base_col+1, pady=(0,6))
            
            nav_matrix.append([])
            nav_matrix.append([])

        for i, stu in enumerate(dm.class_students(self.app.data, cls)):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row_data = {"zh": stu["zh"], "items": {}}
            for it in items:
                pre_sc = _fmt(lt.get("scores", {}).get(stu["zh"], {}).get(it, ""))
                pre_rt = _fmt(lt.get("retakes", {}).get(stu["zh"], {}).get(it, ""))
                row_data["items"][it] = {"score_var": tk.StringVar(value=str(pre_sc)), "retake_var": tk.StringVar(value=str(pre_rt))}
                
            self._rows.append(row_data)

            row = tk.Frame(self._in_inner, bg=bg)
            row.pack(fill="x")
            U.sep(self._in_inner, color=U.C["border"]).pack(fill="x")
            row.columnconfigure(0, minsize=160)
            
            for idx2 in range(len(items)):
                row.columnconfigure(1 + idx2*2,   minsize=70)
                row.columnconfigure(1 + idx2*2+1, minsize=70)
                
            nc = tk.Frame(row, bg=bg)
            nc.grid(row=0, column=0, padx=12, pady=6, sticky="w")
            tk.Label(nc, text=stu["zh"], font=U.FMB, bg=bg, fg=U.C["text"], anchor="w", width=10).pack(anchor="w")
            tk.Label(nc, text=stu.get("en", ""), font=U.FS, bg=bg, fg=U.C["text_lt"], anchor="w", width=10).pack(anchor="w")

            col_counter = 0
            for idx, it in enumerate(items):
                base_col = 1 + idx * 2
                vars_ = row_data["items"][it]

                vcmd_lt = _make_int_score_validator(self) 
                sc_w = tk.Entry(row, textvariable=vars_["score_var"], font=U.FM, width=5,
                                justify="center", relief="flat",
                                highlightbackground=U.C["border"], highlightthickness=1,
                                validate="key", validatecommand=vcmd_lt)
                sc_w.grid(row=0, column=base_col, padx=4, pady=8)
                nav_matrix[col_counter].append(sc_w)
                col_counter += 1

                vcmd_lt2 = _make_int_score_validator(self)
                rt_w = tk.Entry(row, textvariable=vars_["retake_var"], font=U.FM, width=5,
                                justify="center", relief="flat",
                                highlightbackground=U.C["border"], highlightthickness=1,
                                validate="key", validatecommand=vcmd_lt2)
                rt_w.grid(row=0, column=base_col+1, padx=4, pady=8)
                nav_matrix[col_counter].append(rt_w)
                col_counter += 1

        apply_matrix_nav(nav_matrix, canvas=self._in_canvas)

    def _save_all(self):
        cls = self._in_cls_var.get()
        if not cls: 
            messagebox.showinfo("提示", "請先選擇班級", parent=self)
            return
            
        scores = {}
        retakes = {}
        stds = {}
        
        for r in self._rows:
            zh = r["zh"]
            scores[zh] = {}
            retakes[zh] = {}
            stds[zh] = {}
            
            for it, vars_ in r["items"].items():
                sc_s = vars_["score_var"].get().strip()
                rt_s = vars_["retake_var"].get().strip()
                if sc_s:
                    try: 
                        scores[zh][it] = _clean_val(sc_s)
                    except Exception: 
                        scores[zh][it] = sc_s
                if rt_s:
                    try: 
                        retakes[zh][it] = _clean_val(rt_s)
                    except Exception: 
                        retakes[zh][it] = rt_s
                stds[zh][it] = 90

        dm.add_leveltest(self.app.data, cls, self._in_date_var.get(), dm.class_book(self.app.data, cls), scores, retakes, stds)
        self.app.save()
        messagebox.showinfo("完成", "升級考成績已儲存！", parent=self)
        self._show_overview()

# ══════════════════════════════════════════════════════════════════════════════
# 輸出成績單
# ══════════════════════════════════════════════════════════════════════════════
class ExportPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._yr_var = tk.StringVar(value=str(datetime.now().year))
        self._mo_var = tk.StringVar(value=f"{datetime.now().month:02d}")
        self._teacher_var = tk.StringVar() 
        self._check_vars = {}
        self._build()

    def _build(self):
        U.lbl(self, "輸出成績單", size=18, bold=True).pack(anchor="w", padx=20, pady=(14, 8))

        yr_row = tk.Frame(self, bg=U.C["bg"])
        yr_row.pack(fill="x", padx=20, pady=(0, 14))
        U.lbl(yr_row, "年份", size=13, bg=U.C["bg"]).pack(side="left")
        tk.Spinbox(yr_row, textvariable=self._yr_var, from_=2016, to=2036,
                   width=6, font=U.FM, relief="flat", state="readonly",
                   highlightbackground=U.C["border"], highlightthickness=1,
                   buttonbackground=U.C["bg"]).pack(side="left", padx=(4, 12))
        U.lbl(yr_row, "月份", size=13, bg=U.C["bg"]).pack(side="left")
        U.combo(yr_row, self._mo_var, [f"{m:02d}" for m in range(1, 13)], width=5).pack(side="left", padx=(4, 12))
        U.lbl(yr_row, "授課老師", size=13, bg=U.C["bg"]).pack(side="left")
        tk.Entry(yr_row, textvariable=self._teacher_var, font=U.FM, width=12, relief="flat", highlightbackground=U.C["border"], highlightthickness=1).pack(side="left", padx=(4, 12))
        U.btn(yr_row, "套用", self._load_classes, size=12).pack(side="left")

        # 💡 底部操作列 (固定常在)
        bot = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        bot.pack(side="bottom", fill="x", padx=20, pady=(0, 20))
        self._sum_lbl = U.lbl(bot, "", size=11, bg=U.C["white"], fg=U.C["text_lt"])
        self._sum_lbl.pack(side="left", padx=12, pady=7)
        U.btn(bot, "📄 輸出勾選班級成績單", self._export_checked, size=13).pack(side="right", padx=8, pady=5)
        U.ghost_btn(bot, "讀取老師評語", self._fetch_comments, size=13).pack(side="right", padx=4, pady=5)
        U.ghost_btn(bot, "查看缺考名單", self._show_missing_list, size=13).pack(side="right", padx=4, pady=5)

        U.lbl(self, "每月成績單", size=14, bold=True).pack(anchor="w", padx=20, pady=(0, 4))
        
        # ── 月成績區塊 (改為：自動填滿剩餘高度) ──
        U.lbl(self, "選擇要輸出的班級", size=14, bold=True).pack(anchor="w", padx=20, pady=(0, 4))
        self._list_wrap = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        self._list_wrap.pack(fill="both", expand=True, padx=20) # 💡 加入 expand=True 讓它自動長高

        self._exp_header = tk.Frame(self._list_wrap, bg=U.C["header_bg"])
        self._exp_header.pack(fill="x")

        # 💡 移除 height=200 的限制
        self._exp_canvas = tk.Canvas(self._list_wrap, bg=U.C["white"], highlightthickness=0)
        ysb_exp = ttk.Scrollbar(self._list_wrap, orient="vertical", command=self._exp_canvas.yview)
        self._exp_canvas.configure(yscrollcommand=ysb_exp.set)
        ysb_exp.pack(side="right", fill="y")
        self._exp_canvas.pack(side="left", fill="both", expand=True)

        self._rows_frame = tk.Frame(self._exp_canvas, bg=U.C["white"])
        self._exp_canvas.create_window((0, 0), window=self._rows_frame, anchor="nw")
        self._rows_frame.bind("<Configure>", lambda e: self._exp_canvas.configure(scrollregion=self._exp_canvas.bbox("all")))
        U.bind_mousewheel(self._exp_canvas)

        # ── 升級考區塊 (改為：給定固定高度) ──
        U.lbl(self, "升級考成績單", size=14, bold=True).pack(anchor="w", padx=20, pady=(12, 4))

        lt_wrap = tk.Frame(self, bg=U.C["white"], highlightbackground=U.C["border"], highlightthickness=1)
        lt_wrap.pack(fill="x", padx=20, pady=(0, 10)) # 💡 移除 expand=True，讓它固定

        self._lt_header = tk.Frame(lt_wrap, bg=U.C["header_bg"])
        self._lt_header.pack(fill="x")
        
        self.lt_col_defs = [("班級", 0, 100), ("套書", 1, 100), ("考試日期", 2,200), ("待補考", 3, 100), ("老師評語", 4, 120), ("操作", 5, 200)]
        for text, ci, mw in self.lt_col_defs:
            self._lt_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._lt_header, text=text, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=10)

        # 💡 加入 height=250 (這就是升級考區塊的固定高度，覺得太高或太矮可以直接改這個數字)
        self._lt_canvas = tk.Canvas(lt_wrap, bg=U.C["white"], highlightthickness=0, height=100)
        ysb_lt = ttk.Scrollbar(lt_wrap, orient="vertical", command=self._lt_canvas.yview)
        self._lt_canvas.configure(yscrollcommand=ysb_lt.set)
        ysb_lt.pack(side="right", fill="y")
        self._lt_canvas.pack(side="left", fill="both", expand=True)

        self._lt_rows = tk.Frame(self._lt_canvas, bg=U.C["white"])
        self._lt_canvas.create_window((0, 0), window=self._lt_rows, anchor="nw")
        self._lt_rows.bind("<Configure>", lambda e: self._lt_canvas.configure(scrollregion=self._lt_canvas.bbox("all")))
        U.bind_mousewheel(self._lt_canvas)

    def _show_missing_list(self):
        dlg = tk.Toplevel(self)
        dlg.title("本月兒美成績補考名單")
       # dlg.geometry("550x550") #原視窗大小
        dlg.grab_set()
        dlg.update()
        dlg.lift()
        dlg.focus_force()
        
        f = tk.Frame(dlg, bg=U.C["bg"])
        f.pack(fill="both", expand=True)
        
        lbl = U.lbl(f, "本月兒美成績補考名單", size=16, bold=True)
        lbl.pack(pady=(15, 10))
        
        text_area = tk.Text(f, font=U.FM, bg=U.C["white"], relief="flat", highlightbackground=U.C["border"], highlightthickness=1, padx=10, pady=10)
        text_area.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        output_lines = []
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        
        for cls in dm.class_list(self.app.data):
            cls_name = cls if "班" in cls else f"{cls}班"
            missing_info = []

            all_recs = dm.all_monthly_records(self.app.data, cls, ym)
            class_test_keys = set()
            for recs in all_recs.values():
                for r in recs:
                    class_test_keys.add((r.get("type"), r.get("date"), r.get("range"), r.get("item")))
            
            sorted_keys = sorted(list(class_test_keys), key=lambda x: x[1]) 
            
            for test_key in sorted_keys:
                t_type, t_date, t_range, t_item = test_key
                is_speaking = (t_type == "口說")
                missing_stus = []
                for stu in dm.class_students(self.app.data, cls):
                    stu_recs = all_recs.get(stu["zh"], [])
                    rec = next((r for r in stu_recs if (r.get("type"), r.get("date"), r.get("range"), r.get("item")) == test_key), None)
                    
                    is_missing = False
                    if not rec:
                        is_missing = True
                    else:
                        sc = str(rec.get("score", ""))
                        rt = str(rec.get("retake", ""))
                        std = rec.get("std", stu.get("std", 90))
                        
                        if not sc:
                            is_missing = True
                        elif not is_speaking:
                            def _ok(v, s):
                                if not v: return False
                                try: return float(v) >= float(s)
                                except ValueError: return str(v) in ("A++", "A+", "A")
                            if not _ok(sc, std) and not _ok(rt, std):
                                is_missing = True
                                
                    if is_missing:
                        name = stu.get("en", "").strip() or stu.get("zh", "").strip()
                        missing_stus.append(name)
                
                if missing_stus:
                    date_str = t_date[5:].replace("-", "/") if len(t_date) >= 5 else t_date
                    ri_str = f"{t_range} {t_item}".strip()
                    type_prefix = "" if t_type == "考試本" else f"[{t_type}] "
                    missing_info.append(f"{date_str} {type_prefix}{ri_str} {'、'.join(missing_stus)}".replace("  ", " "))

            lts = self.app.data.get("classes", {}).get(cls, {}).get("leveltests", [])
            if not lts:
                lts = self.app.data.get("leveltest", {}).get(cls, [])
                
            for lt in sorted(lts, key=lambda x: x.get("date", "")):
                lt_date = lt.get("date", "")
                ym_from_date = lt_date[:7].replace("/", "-") if "/" in lt_date else lt_date[:7]
                if ym_from_date != ym: continue 

                book = lt.get("book", "")
                scores = lt.get("scores", {})
                retakes = lt.get("retakes", {})
                stds = lt.get("stds", {})
                
                all_items = set()
                for zh_scores in scores.values():
                    all_items.update(zh_scores.keys())
                if not all_items:
                    all_items = set(dm.book_items(self.app.data, book))
                
                for it in sorted(list(all_items)):
                    missing_stus = []
                    for stu in dm.class_students(self.app.data, cls):
                        zh = stu["zh"]
                        sc = scores.get(zh, {}).get(it, "")
                        rt = retakes.get(zh, {}).get(it, "")
                        std = stds.get(zh, {}).get(it, 90)
                        
                        def _ok(v, s):
                            if v == "" or v is None: return False
                            try: return float(v) >= float(s)
                            except ValueError: return str(v) in ("A++", "A+", "A")
                        
                        if not _ok(sc, std) and not _ok(rt, std):
                            name = stu.get("en", "").strip() or stu.get("zh", "").strip()
                            missing_stus.append(name)
                            
                    if missing_stus:
                        date_str = lt_date[5:].replace("-", "/") if len(lt_date) >= 5 else lt_date
                        missing_info.append(f"{date_str} [升級考] {book}-{it} {'、'.join(missing_stus)}")

            if missing_info:
                output_lines.append(f"{cls_name}: \n    " + "\n    ".join(missing_info))
            else:
                output_lines.append(f"{cls_name}: 無待處理紀錄")
        
# 💡 將文字組合起來
        final_text = "\n\n".join(output_lines)
        text_area.insert("1.0", final_text)
        
        # 💡 動態計算行數 (最高不超過 25 行，避免內容太多時視窗超出螢幕)
        total_lines = final_text.count("\n") + 2
        display_height = max(5, min(total_lines, 25))
        text_area.config(state="disabled", height=display_height, width=55)

    def refresh(self):
        self._load_classes()
        self._render_lt_rows()

    def _load_classes(self):
        for w in self._exp_header.winfo_children(): w.destroy()
        for w in self._rows_frame.winfo_children(): w.destroy()
        self._check_vars = {}
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        
        col_defs = [("", 0, 60), ("班級", 1, 200), ("缺考/補考", 2, 200), ("老師評語", 3, 200), ("", 4, 300)]
        for text, ci, mw in col_defs:
            self._exp_header.columnconfigure(ci, minsize=mw)
            tk.Label(self._exp_header, text=text, font=U.FMB, bg=U.C["header_bg"], fg=U.C["text_lt"], anchor="w", pady=6).grid(row=0, column=ci, sticky="w", padx=10)

        for i, cls in enumerate(dm.class_list(self.app.data)):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            
            row = tk.Frame(self._rows_frame, bg=bg)
            row.pack(fill="x")
            U.sep(self._rows_frame, color=U.C["border"]).pack(fill="x")
            for _, ci, mw in col_defs: 
                row.columnconfigure(ci, minsize=mw)
            
            v = tk.BooleanVar(value=True)
            self._check_vars[cls] = v
            v.trace_add("write", lambda *a: self._update_summary())
            
            cb = tk.Checkbutton(row, variable=v, bg=bg)
            cb.grid(row=0, column=0, sticky="w", padx=10, pady=6)
            
            tk.Label(row, text=cls, font=U.FMB, bg=bg, fg=U.C["text"]).grid(row=0, column=1, sticky="w", padx=10, pady=6)
            
            all_recs = dm.all_monthly_records(self.app.data, cls, ym)
            class_test_keys = set()
            for recs in all_recs.values():
                for r in recs:
                    class_test_keys.add((r.get("type"), r.get("date"), r.get("range"), r.get("item")))
            
            missing = 0
            for stu in dm.class_students(self.app.data, cls):
                stu_recs = all_recs.get(stu["zh"], [])
                for test_key in class_test_keys:
                    t_type, _, _, _ = test_key
                    is_speaking = (t_type == "口說")
                    rec = next((r for r in stu_recs if (r.get("type"), r.get("date"), r.get("range"), r.get("item")) == test_key), None)
                    
                    if not rec:
                        missing += 1
                    else:
                        sc = str(rec.get("score", ""))
                        rt = str(rec.get("retake", ""))
                        std = _fmt(stu.get("std", 90))
                        
                        if not sc:
                            missing += 1
                        elif not is_speaking:
                            def _ok(v, s):
                                if not v: return False
                                try: return float(v) >= float(s)
                                except ValueError: return str(v) in ("A++", "A+", "A")
                            if not _ok(sc, std) and not _ok(rt, std):
                                missing += 1

            pc = U.C["ng_fg"] if missing > 0 else U.C["text_lt"]
            tk.Label(row, text=f"{missing} 筆" if missing > 0 else "—", font=U.FMB if missing > 0 else U.FM, bg=bg, fg=pc).grid(row=0, column=2, sticky="w", padx=10, pady=6)
            
            # 評語填寫狀態
            comments_data = self.app.data.get("comments", {}).get(cls, {}).get(ym, {})
            stu_list = dm.class_students(self.app.data, cls)
            filled_cnt = sum(1 for s in stu_list if comments_data.get(s["zh"], "").strip())
            total_cnt = len(stu_list)
            if filled_cnt == total_cnt and total_cnt > 0:
                cmt_text = "✓ 已完成"
                cmt_color = U.C["ok_fg"] if hasattr(U.C, "ok_fg") else "#27ae60"
            elif filled_cnt > 0:
                cmt_text = f"{filled_cnt}/{total_cnt} 份"
                cmt_color = U.C["ng_fg"]
            else:
                cmt_text = "未填寫"
                cmt_color = U.C["text_lt"]
            tk.Label(row, text=cmt_text, font=U.FM, bg=bg, fg=cmt_color).grid(row=0, column=3, sticky="w", padx=10, pady=6)

            op = tk.Frame(row, bg=bg)
            op.grid(row=0, column=4, sticky="w", padx=10, pady=6)
            U.ghost_btn(op, "單獨輸出", lambda c=cls: self._export_one(c), size=12).pack(side="left")

        self._update_summary()

    def _update_summary(self):
        checked = [c for c, v in self._check_vars.items() if v.get()]
        self._sum_lbl.config(text=f"已勾選 {len(checked)} 個班級　共 {sum(len(dm.class_students(self.app.data, c)) for c in checked)} 份成績單")

    def _get_template(self):
        import sys; from pathlib import Path
        base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
        return next((str(p) for p in [base / "成績單模板.docx"] if p.exists()), None)

    def _render_lt_rows(self):
        import data_manager as dm
        for w in self._lt_rows.winfo_children(): 
            w.destroy()

        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        records = []
        for cls, cdata in self.app.data.get("classes", {}).items():
            lts = cdata.get("leveltests", [])
            if not lts:
                lts = self.app.data.get("leveltest", {}).get(cls, [])
            for lt in lts:
                records.append({"cls": cls, "lt": lt})

        if not records:
            U.lbl(self._lt_rows, "目前尚無升級考紀錄", size=14,
                  fg=U.C["text_lt"], bg=U.C["white"]).pack(padx=12, pady=8)
            return

        for i, rec in enumerate(records):
            cls = rec["cls"]
            lt = rec["lt"]
            date = lt.get("date", "")
            book = lt.get("book", "")
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]

            scores  = lt.get("scores", {})
            retakes = lt.get("retakes", {})
            stds    = lt.get("stds", {})
            
            all_items = set()
            for zh_scores in scores.values():
                all_items.update(zh_scores.keys())
            if not all_items:
                all_items = set(dm.book_items(self.app.data, book))
                
            pending = 0
            for stu in dm.class_students(self.app.data, cls):
                zh = stu["zh"]
                for it in all_items:
                    sc = scores.get(zh, {}).get(it, "")
                    rt = retakes.get(zh, {}).get(it, "")
                    std = _fmt(stu.get("std", 90))
                    def _ok(v, s):
                        if not v: return False
                        try: return float(v) >= float(s)
                        except ValueError: return str(v) in ("A++", "A+", "A")
                    
                    if not _ok(sc, std) and not _ok(rt, std):
                        pending += 1

            row = tk.Frame(self._lt_rows, bg=bg)
            row.pack(fill="x")
            for _, ci, mw in self.lt_col_defs:
                row.columnconfigure(ci, minsize=mw)

            tk.Label(row, text=cls, font=U.FMB, bg=bg, fg=U.C["primary"]).grid(row=0, column=0, sticky="w", padx=10, pady=6)
            tk.Label(row, text=book, font=U.FM, bg=bg, fg=U.C["text_lt"]).grid(row=0, column=1, sticky="w", padx=10, pady=6)
            tk.Label(row, text=date, font=U.FM, bg=bg, fg=U.C["text"]).grid(row=0, column=2, sticky="w", padx=10, pady=6)
            
            pc = U.C["ng_fg"] if pending > 0 else U.C["text_lt"]
            tk.Label(row, text=f"{pending} 筆" if pending > 0 else "—", font=U.FMB if pending > 0 else U.FM, bg=bg, fg=pc).grid(row=0, column=3, sticky="w", padx=10, pady=6)

            # 升級考評語狀態（從 lt_comments 讀）
            lt_comments_data = self.app.data.get("lt_comments", {}).get(cls, {}).get(date, {})
            stu_list = dm.class_students(self.app.data, cls)
            lt_filled = sum(1 for s in stu_list if lt_comments_data.get(s["zh"], {}).get("comment", "").strip())
            lt_total = len(stu_list)
            if lt_filled == lt_total and lt_total > 0:
                lt_cmt_text = "✓ 已完成"
                lt_cmt_color = "#27ae60"
            elif lt_filled > 0:
                lt_cmt_text = f"{lt_filled}/{lt_total} 份"
                lt_cmt_color = U.C["ng_fg"]
            else:
                lt_cmt_text = "未填寫"
                lt_cmt_color = U.C["text_lt"]
            tk.Label(row, text=lt_cmt_text, font=U.FM, bg=bg, fg=lt_cmt_color).grid(row=0, column=4, sticky="w", padx=10, pady=6)

            op = tk.Frame(row, bg=bg)
            op.grid(row=0, column=5, sticky="w", padx=10, pady=6)
            U.ghost_btn(op, "📄 輸出", lambda c=cls, d=date: self._export_lt(c, d), size=13).pack(side="left")

    def _export_lt(self, cls, date):
        import docx_exporter as dx, sys, os
        import data_manager as dm
        from pathlib import Path

        lts = self.app.data["classes"][cls].get("leveltests", [])
        if not lts:
            lts = self.app.data.get("leveltest", {}).get(cls, [])
        lt = next((x for x in lts if x.get("date") == date), None)
        if not lt:
            messagebox.showwarning("找不到資料", f"找不到 {cls} 於 {date} 的升級考紀錄", parent=self)
            return

        default_lt = self.app.data.get("settings", {}).get("lt_dir", "")
        folder = filedialog.askdirectory(title="選擇儲存資料夾", parent=self,
                                         initialdir=default_lt or None)

        # 用書本設定的 items 順序（聽力、單字、翻譯、口說）
        all_items = dm.book_items(self.app.data, lt.get("book", ""))
        if not all_items:
            # fallback：從成績資料收集
            all_items_set = set()
            for zh_scores in lt.get("scores", {}).values():
                all_items_set.update(zh_scores.keys())
            all_items = sorted(list(all_items_set))

        records_for_export = []
        for zh in dm.student_zh_list(self.app.data, cls):
            stu_scores = lt.get("scores", {}).get(zh, {})
            items_data = {it: {
                "score":  stu_scores.get(it, ""),
                "retake": lt.get("retakes", {}).get(zh, {}).get(it, ""),
                "std":    _fmt(next((s.get("std", 90) for s in dm.class_students(self.app.data, cls) if s["zh"] == zh), 90)),
            } for it in all_items}
            records_for_export.append({"zh": zh, "items": items_data})

        base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
        tmpl = next((str(p) for p in [base / "升級考模板.docx"] if p.exists()), None)

        ym_from_date = date[:7].replace("/", "-") if "/" in date else date[:7]
        lt_comments = self.app.data.get("lt_comments", {}).get(cls, {}).get(date, {})
        teacher = dm.class_teacher(self.app.data, cls)
        sig_key_lt = teacher + '_' + cls + '_lt'
        sig_key    = teacher + '_' + cls
        sig_b64 = self.app.data.get("signatures", {}).get(sig_key_lt,
                  self.app.data.get("signatures", {}).get(sig_key,
                  self.app.data.get("signatures", {}).get(teacher, None)))

        try:
            filepath = dx.export_leveltest(folder, cls, lt["book"], date,
                                           dm.class_students(self.app.data, cls),
                                           records_for_export, tmpl,
                                           lt_comments=lt_comments, sig_b64=sig_b64)
            if filepath:
                messagebox.showinfo("完成", "已輸出升級考成績單！\n" + filepath, parent=self)
                try: os.startfile(filepath)
                except Exception: pass
        except PermissionError:
            messagebox.showwarning("檔案被占用",
                "同名檔案已開啟中，請先關閉後再重新輸出。", parent=self)
        except Exception as e:
            messagebox.showerror("輸出失敗", f"匯出成績單時發生錯誤：\n\n{e}", parent=self)

    def _get_comments_goals_sig(self, cls, ym):
        comments  = self.app.data.get('comments', {}).get(cls, {}).get(ym, {})
        goals     = self.app.data.get('goals', {}).get(cls, {}).get(ym, [])
        teacher   = dm.class_teacher(self.app.data, cls)
        sig_key = teacher + '_' + cls
        sig_b64 = self.app.data.get('signatures', {}).get(sig_key,
                  self.app.data.get('signatures', {}).get(teacher, None))
        return comments, goals, sig_b64

    def _export_one(self, cls):
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        teacher_name = dm.class_teacher(self.app.data, cls)
        default_dir = self.app.data.get("settings", {}).get("monthly_dir", "")
        folder = filedialog.askdirectory(title="選擇儲存資料夾", parent=self,
                                         initialdir=default_dir or None)
        if not folder: return
        import docx_exporter as dx
        students = dm.class_students(self.app.data, cls)
        book = dm.class_book(self.app.data, cls)
        
        students_data = [{"stu": stu, "records": dm.monthly_records(self.app.data, cls, stu["zh"], ym)} for stu in students]
        comments, goals, sig_b64 = self._get_comments_goals_sig(cls, ym)

        try:
            path = dx.export_monthly_batch(
                folder=folder,
                template_path=self._get_template(),
                cls_name=cls,
                book_name=book,
                teacher=teacher_name,
                year_month=ym,
                students_data=students_data,
                comments=comments,
                goals=goals,
                sig_b64=sig_b64,
            )
            if path:
                messagebox.showinfo("完成", f"已輸出：\n{path}", parent=self)
                import os
                try: os.startfile(path)
                except Exception: pass
        except PermissionError:
            messagebox.showwarning("檔案被占用",
                "同名檔案已開啟中，請先關閉後再重新輸出。", parent=self)
        except Exception as e:
            messagebox.showerror("輸出失敗", f"匯出成績單時發生錯誤：\n\n{e}", parent=self)

    def _export_checked(self):
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        checked = [c for c, v in self._check_vars.items() if v.get()]
        if not checked: return
        default_batch = self.app.data.get("settings", {}).get("monthly_dir", "")
        folder = filedialog.askdirectory(title="選擇儲存資料夾", parent=self,
                                         initialdir=default_batch or None)
        
        import docx_exporter as dx
        try:
            for cls in checked:
                teacher_name = dm.class_teacher(self.app.data, cls) 
                students = dm.class_students(self.app.data, cls)
                book = dm.class_book(self.app.data, cls)
                
                students_data = [{"stu": stu, "records": dm.monthly_records(self.app.data, cls, stu["zh"], ym)} for stu in students]
                
                comments, goals, sig_b64 = self._get_comments_goals_sig(cls, ym)
                dx.export_monthly_batch(
                    folder=folder,
                    template_path=self._get_template(),
                    cls_name=cls,
                    book_name=book,
                    teacher=teacher_name,
                    year_month=ym,
                    students_data=students_data,
                    comments=comments,
                    goals=goals,
                    sig_b64=sig_b64,
                )
            messagebox.showinfo("完成", f"已輸出 {len(checked)} 個班級的成績單\n儲存於：{folder}", parent=self)
            import os
            try: os.startfile(folder)
            except Exception: pass
        except PermissionError:
            messagebox.showwarning("檔案被占用",
                "同名檔案已開啟中，請先關閉後再重新輸出。", parent=self)
        except Exception as e:
             messagebox.showerror("輸出失敗", f"匯出成績單時發生錯誤：\n\n{e}", parent=self)

    def _fetch_comments(self):
        import urllib.request, json, threading

        SHEET_URL = self.app.data.get("settings", {}).get(
            "gs_url",
            'https://script.google.com/macros/s/AKfycbzKgnCmcBH_CJcdtZx_XkZ07nvfKPc9Y-xGUn-ahgXMKcSeCVBBwZKAmjexxYsNBOW8/exec'
        )

        # 非阻塞讀取視窗
        loading_win = tk.Toplevel(self)
        loading_win.title("讀取中")
        loading_win.geometry("280x80")
        loading_win.resizable(False, False)
        loading_win.grab_set()
        tk.Label(loading_win, text="正在從雲端讀取評語，請稍候...",
                 font=U.FM, pady=20).pack()
        loading_win.update()

        def _do_fetch():
            try:
                import ssl, urllib.parse
                ctx = ssl.create_default_context()
                opener = urllib.request.build_opener(
                    urllib.request.HTTPSHandler(context=ctx),
                    urllib.request.HTTPCookieProcessor()
                )
                req = urllib.request.Request(
                    SHEET_URL,
                    headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                        'Accept': 'application/json, text/javascript, */*',
                    }
                )
                with opener.open(req, timeout=20) as resp:
                    raw = resp.read().decode('utf-8')
                # Apps Script 有時回傳 HTML（重導向）不是 JSON
                if not raw.strip().startswith('{'):
                    raise Exception("回應不是 JSON，可能需要重新登入 Google 帳號或重新部署")
                result = json.loads(raw)
                if result.get('status') != 'ok':
                    self.after(0, lambda: (loading_win.destroy(),
                        messagebox.showerror("錯誤", "Google Sheet 回應異常：" + str(result), parent=self)))
                    return
                self.after(0, lambda: (loading_win.destroy(), self._apply_comments(result)))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: (loading_win.destroy(),
                    messagebox.showerror("讀取失敗",
                        "錯誤訊息：" + err, parent=self)))

        threading.Thread(target=_do_fetch, daemon=True).start()

    def _apply_comments(self, result):
        """result 是 doGet 回傳的整個 dict，含 monthly 和 leveltest"""
        ym = f"{self._yr_var.get()}-{self._mo_var.get()}"
        updated = 0
        sig_cache = {}

        # 讀月成績單評語
        self._apply_monthly(result.get('monthly', []), ym, sig_cache)
        updated += self._apply_lt_comments(result.get('leveltest', []))

        self.app.save()
        self._load_classes()
        self._render_lt_rows()
        # 統計升級考簽名
        lt_sigs = list(set(
            k.rsplit('_lt',1)[0]
            for k in self.app.data.get('signatures',{})
            if k.endswith('_lt')
        ))
        monthly_sigs = list(set(
            k for k in sig_cache
        ))
        total = updated + len(sig_cache)

        if total > 0 or lt_sigs:
            parts = []
            if sig_cache: parts.append(f"月成績單：{'、'.join(set(k.split('_')[0] for k in sig_cache))}")
            if lt_sigs:   parts.append(f"升級考：{'、'.join(lt_sigs)}")
            sig_str = '\n'.join(parts) if parts else '無'
            messagebox.showinfo("完成",
                f"已讀取評語資料\n月份：{ym}\n"
                f"升級考評語：{updated} 筆\n"
                f"簽名：{sig_str}\n\n"
                "下次輸出成績單時將自動帶入評語。", parent=self)
        else:
            messagebox.showinfo("提示",
                f"找不到 {ym} 的評語資料\n請確認老師已填寫並送出表單。", parent=self)

    def _apply_monthly(self, rows, ym, sig_cache):
        """處理月成績單評語"""
        import json as _json
        for row in rows:
            month_raw = str(row.get('月份', '')).strip()
            month     = month_raw[:7] if len(month_raw) >= 7 else month_raw
            cls       = str(row.get('班級', '')).strip()
            zh        = str(row.get('學生姓名', '')).strip()
            comment   = str(row.get('評語', '')).strip()
            teacher   = str(row.get('老師', '')).strip()
            sig       = str(row.get('簽名圖(base64)', '')).strip()
            status    = str(row.get('狀態', '')).strip()
            goals_raw = [str(row.get(f'教學目標{i}', '')).strip() for i in range(1, 4)]
            goals     = [g for g in goals_raw if g]

            if status == '草稿': continue
            if month != ym: continue
            if not cls or not zh: continue

            self.app.data.setdefault('comments', {})                 .setdefault(cls, {})                 .setdefault(ym, {})[zh] = comment

            if goals:
                self.app.data.setdefault('goals', {})                     .setdefault(cls, {})[ym] = goals

            sig_key = teacher + '_' + cls
            if sig and sig_key not in sig_cache:
                sig_cache[sig_key] = sig
                self.app.data.setdefault('signatures', {})[sig_key] = sig

    def _apply_lt_comments(self, rows):
        """處理升級考評語"""
        import json as _json
        updated = 0
        for row in rows:
            cls      = str(row.get('班級', '')).strip()
            zh       = str(row.get('學生姓名', '')).strip()
            comment  = str(row.get('評語', '')).strip()
            teacher  = str(row.get('老師', '')).strip()
            sig      = str(row.get('簽名圖(base64)', '')).strip()
            lt_date_raw = str(row.get('考試日期', '') or '').strip()
            # Google Sheets 把日期轉成 ISO 格式（UTC），需加8小時轉台灣時間
            if 'T' in lt_date_raw and 'Z' in lt_date_raw:
                try:
                    from datetime import datetime, timedelta
                    dt_utc = datetime.fromisoformat(lt_date_raw.replace('Z', '+00:00'))
                    dt_tw = dt_utc + timedelta(hours=8)
                    lt_date = dt_tw.strftime('%Y-%m-%d')
                except Exception:
                    lt_date = lt_date_raw[:10]
            else:
                lt_date = lt_date_raw
            status   = str(row.get('狀態', '')).strip()
            lt_checks_raw = str(row.get('升級考勾選', '') or '').strip()
            lt_result     = str(row.get('評鑑結果',   '') or '').strip()

            if status == '草稿': continue
            if not cls or not zh or not lt_date: continue

            try: _checks = _json.loads(lt_checks_raw)
            except: _checks = []

            lt_date_norm = lt_date.replace('-', '/')
            self.app.data.setdefault('lt_comments', {})                 .setdefault(cls, {})                 .setdefault(lt_date_norm, {})                 [zh] = {'comment': comment, 'checks': _checks, 'result': lt_result}

            sig_key = teacher + '_' + cls + '_lt'
            if sig:
                self.app.data.setdefault('signatures', {})[sig_key] = sig

            updated += 1
        return updated




# ══════════════════════════════════════════════════════════════════════════════
# 系統設定
# ══════════════════════════════════════════════════════════════════════════════
class SettingsPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=U.C["bg"])
        self.app = app
        self._build()

    def _build(self):
        # 標題
        topbar = tk.Frame(self, bg=U.C["bg"])
        topbar.pack(fill="x", padx=20, pady=(12, 6))
        U.lbl(topbar, "系統設定", size=16, bold=True, bg=U.C["bg"]).pack(side="left")

        # 可捲動的內容區
        wrap = tk.Frame(self, bg=U.C["bg"])
        wrap.pack(fill="both", expand=True)
        canvas = tk.Canvas(wrap, bg=U.C["bg"], highlightthickness=0)
        ysb = ttk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=U.C["bg"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        def _on_inner_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_resize(e):
            canvas.itemconfig(win_id, width=e.width)
        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_resize)
        U.bind_mousewheel(canvas)

        body = tk.Frame(inner, bg=U.C["bg"])
        body.pack(fill="x", padx=20, pady=(0, 16))

        def make_sec(title, add_cmd=None, add_label=None):
            """建立區塊，回傳內容 frame"""
            sec = tk.Frame(body, bg=U.C["white"],
                           highlightbackground=U.C["border"], highlightthickness=1)
            sec.pack(fill="x", pady=(0, 10))
            hdr = tk.Frame(sec, bg=U.C["header_bg"])
            hdr.pack(fill="x")
            U.lbl(hdr, title, size=11, bold=True, bg=U.C["header_bg"],
                  fg=U.C["text_lt"]).pack(side="left", padx=10, pady=5)
            if add_cmd:
                U.btn(hdr, add_label or "＋ 新增", add_cmd, size=10).pack(
                    side="right", padx=6, pady=3)
            f = tk.Frame(sec, bg=U.C["white"])
            f.pack(fill="x", padx=12, pady=6)
            return f

        def make_field(parent, label, var, width=None, hint=None, browse_cmd=None, expand=False):
            """單欄設定列
            expand=True 或有 browse_cmd：Entry 填滿剩餘寬度
            否則：Entry 固定寬度（width 參數）
            """
            row = tk.Frame(parent, bg=U.C["white"])
            row.pack(fill="x", pady=2)

            full_width = expand or bool(browse_cmd)
            if full_width:
                row.columnconfigure(1, weight=1)

            U.lbl(row, label, size=10, bg=U.C["white"], fg=U.C["text_lt"],
                  anchor="w", width=16).grid(row=0, column=0, sticky="w", padx=(0,6))

            kw = dict(textvariable=var, font=U.FM, relief="flat",
                      highlightbackground=U.C["border"], highlightthickness=1)
            if full_width:
                e = tk.Entry(row, **kw)
                e.grid(row=0, column=1, sticky="ew", padx=(0,4))
            else:
                e = tk.Entry(row, width=width or 8, **kw)
                e.grid(row=0, column=1, sticky="w", padx=(0,4))

            col = 2
            if hint:
                U.lbl(row, hint, size=9, bg=U.C["white"],
                      fg=U.C["text_lt"]).grid(row=0, column=col, sticky="w")
                col += 1
            if browse_cmd:
                U.ghost_btn(row, "瀏覽...", browse_cmd, size=9).grid(
                    row=0, column=col, sticky="w", padx=(4, 0))
            return e

        cfg = self.app.data.setdefault("settings", {})

        # ── 老師管理 ──────────────────────────────────────────────────────────
        self._teacher_list_frame = make_sec("老師管理", self._add_teacher, "＋ 新增老師")

        # ── 版面設定 ──────────────────────────────────────────────────────────
        f_layout = make_sec("版面設定")
        dm_row = tk.Frame(f_layout, bg=U.C["white"])
        dm_row.pack(fill="x", pady=2)
        U.lbl(dm_row, "深色模式", size=10, bg=U.C["white"], fg=U.C["text_lt"],
              anchor="w", width=16).pack(side="left", padx=(0,6))
        self._dark_var = tk.BooleanVar(value=cfg.get("dark_mode", False))
        tk.Checkbutton(dm_row, variable=self._dark_var, font=U.FM,
                       bg=U.C["white"], activebackground=U.C["white"],
                       cursor="hand2", text="開啟深色模式",
                       command=self._toggle_dark).pack(side="left")

        # ── 成績單設定 ────────────────────────────────────────────────────────
        f2 = make_sec("成績單設定")

        self._gs_url_var = tk.StringVar(value=cfg.get("gs_url", ""))
        e_gs = make_field(f2, "Google Sheet 網址", self._gs_url_var, expand=True)
        e_gs.bind("<FocusOut>", lambda e: self._save_cfg("gs_url", self._gs_url_var))
        e_gs.bind("<Return>",   lambda e: self._save_cfg("gs_url", self._gs_url_var))

        U.sep(f2, color=U.C["border"]).pack(fill="x", pady=4)

        default_std = self.app.data.get("settings", {}).get("default_std", 90)
        self._std_var = tk.StringVar(value=str(default_std))
        e_std = make_field(f2, "基本標準分", self._std_var, width=6, hint="分（全班預設值）")
        e_std.bind("<FocusOut>", lambda e: self._save_cfg("default_std", self._std_var, int))
        e_std.bind("<Return>",   lambda e: self._save_cfg("default_std", self._std_var, int))

        U.sep(f2, color=U.C["border"]).pack(fill="x", pady=4)

        self._monthly_dir_var = tk.StringVar(value=cfg.get("monthly_dir", ""))
        make_field(f2, "月成績單輸出資料夾", self._monthly_dir_var, width=28,
                   browse_cmd=lambda: self._browse_dir("monthly_dir", self._monthly_dir_var))

        self._lt_dir_var = tk.StringVar(value=cfg.get("lt_dir", ""))
        make_field(f2, "升級考成績單資料夾", self._lt_dir_var, width=28,
                   browse_cmd=lambda: self._browse_dir("lt_dir", self._lt_dir_var))



    def refresh(self):
        self._render_teachers()
        # 同步設定值到欄位
        cfg = self.app.data.get("settings", {})
        if hasattr(self, '_gs_url_var'):
            self._gs_url_var.set(cfg.get("gs_url", ""))
            self._std_var.set(str(cfg.get("default_std", 90)))
            self._monthly_dir_var.set(cfg.get("monthly_dir", ""))
            self._lt_dir_var.set(cfg.get("lt_dir", ""))
        if hasattr(self, '_dark_var'):
            self._dark_var.set(cfg.get("dark_mode", False))

    def _apply_font_size(self):
        """儲存並套用字體大小"""
        size_map = {"小": 9, "中": 10, "大": 12}
        sz = self._font_size_var.get()
        self._save_cfg("font_size", self._font_size_var)
        pts = size_map.get(sz, 10)
        # 更新 ui_utils 的字型（下次開啟的視窗生效）
        U.FM = (U.F, pts)
        U.FMB = (U.F, pts, "bold")
        U.FL = (U.F, pts + 1)
        U.FLB = (U.F, pts + 1, "bold")
        messagebox.showinfo("字體大小", "已設定為「" + sz + "」\n重新啟動程式後完全生效。", parent=self)

    def _toggle_dark(self):
        """深色模式切換並立即套用"""
        val = self._dark_var.get()
        self.app.data.setdefault("settings", {})["dark_mode"] = val
        self.app.save()
        U.apply_theme(val)
        mode_str = "開啟" if val else "關閉"
        messagebox.showinfo("深色模式",
            "已" + mode_str + "深色模式\n重新啟動程式後完全生效。",
            parent=self)
    def _save_cfg(self, key, var, cast=None):
        val = var.get().strip()
        if cast:
            try: val = cast(val)
            except: return
        self.app.data.setdefault("settings", {})[key] = val
        self.app.save()

    def _browse_dir(self, key, var):
        """瀏覽資料夾"""
        from tkinter import filedialog
        path = filedialog.askdirectory(title="選擇資料夾", parent=self)
        if path:
            var.set(path)
            self._save_cfg(key, var)

    def _render_teachers(self):
        for w in self._teacher_list_frame.winfo_children():
            w.destroy()

        teachers = dm.teacher_list(self.app.data)
        if not teachers:
            U.lbl(self._teacher_list_frame, "尚未設定老師，請點「＋ 新增老師」",
                  size=12, fg=U.C["text_lt"]).pack(pady=8)
            return

        for i, teacher in enumerate(teachers):
            bg = U.C["white"] if i % 2 == 0 else U.C["row_alt"]
            row = tk.Frame(self._teacher_list_frame, bg=bg)
            row.pack(fill="x")
            U.sep(self._teacher_list_frame, color=U.C["border"]).pack(fill="x")

            U.lbl(row, teacher, size=11, bold=False, bg=bg).pack(
                side="left", padx=12, pady=5)

            op = tk.Frame(row, bg=bg)
            op.pack(side="right", padx=8, pady=4)
            U.ghost_btn(op, "改名", lambda t=teacher: self._rename_teacher(t),
                        size=12).pack(side="left", padx=3)
            U.danger_btn(op, "刪除", lambda t=teacher: self._delete_teacher(t),
                         size=12).pack(side="left", padx=3)

    def _add_teacher(self):
        name = U.ask_string(self, "新增老師", "老師姓名：")
        if not name:
            return
        if name in dm.teacher_list(self.app.data):
            messagebox.showwarning("重複", f"「{name}」已存在", parent=self)
            return
        dm.add_teacher(self.app.data, name)
        self.app.save()
        self._render_teachers()
        # 通知班級頁面更新下拉選單
        if hasattr(self.app, '_pages') and 'classes' in self.app._pages:
            self.app._pages['classes'].refresh()

    def _rename_teacher(self, old_name):
        new_name = U.ask_string(self, "改名", "新名稱：", default=old_name)
        if not new_name or new_name == old_name:
            return
        if new_name in dm.teacher_list(self.app.data):
            messagebox.showwarning("重複", f"「{new_name}」已存在", parent=self)
            return
        # 更新老師名單
        teachers = self.app.data.get("teachers", [])
        idx = teachers.index(old_name)
        teachers[idx] = new_name
        # 更新各班班級的老師設定
        for cls_data in self.app.data.get("classes", {}).values():
            if cls_data.get("teacher") == old_name:
                cls_data["teacher"] = new_name
        self.app.save()
        self._render_teachers()
        if hasattr(self.app, '_pages') and 'classes' in self.app._pages:
            self.app._pages['classes'].refresh()

    def _delete_teacher(self, name):
        # 確認有無班級使用此老師
        using = [c for c, d in self.app.data.get("classes", {}).items()
                 if d.get("teacher") == name]
        msg = f"確定刪除老師「{name}」？"
        if using:
            msg += f"\n\n以下班級的授課老師將被清空：\n{'、'.join(using)}"
        if not messagebox.askyesno("確認刪除", msg, parent=self):
            return
        dm.remove_teacher(self.app.data, name)
        self.app.save()
        self._render_teachers()
        if hasattr(self.app, '_pages') and 'classes' in self.app._pages:
            self.app._pages['classes'].refresh()
