"""
excel_importer.py
將現有的兒美成績輸入表.xlsx 匯入到系統資料庫
執行方式：把此檔放在程式資料夾，雙擊「匯入舊Excel資料.bat」
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

# 月份中文對照
MONTH_MAP = {
    '一': '01', '二': '02', '三': '03', '四': '04',
    '五': '05', '六': '06', '七': '07', '八': '08',
    '九': '09', '十': '10', '十一': '11', '十二': '12',
}

# 課程對應套書名稱
COURSE_BOOK_MAP = {
    '發音1': '發音1', '發音2': '發音2', '發音3': '發音3', '發音4': '發音4',
    '文法1': '文法1', '文法2': '文法2', '文法3': '文法3', '文法4': '文法4',
}

import os

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"

DATA_FILE = DATA_DIR / "grades.json"
DATA_DIR.mkdir(parents=True, exist_ok=True)


def load_data():
    if DATA_FILE.exists():
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {
        "books": [
            {"name": "發音1", "items": ["聽力", "字母", "口說"]},
            {"name": "發音2", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "發音3", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "發音4", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "文法1", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "文法2", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "文法3", "items": ["聽力", "單字", "翻譯", "口說"]},
            {"name": "文法4", "items": ["聽力", "單字", "翻譯", "口說"]},
        ],
        "classes": {},
        "monthly": {},
        "leveltest": {},
        "upgraded": {},
    }


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def parse_name(name_str):
    """解析 '王小明(Tommy)' → (zh='王小明', en='Tommy')"""
    name_str = str(name_str).strip()
    if '(' in name_str and ')' in name_str:
        zh = name_str[:name_str.index('(')].strip()
        en = name_str[name_str.index('(')+1:name_str.index(')')].strip()
    else:
        zh = name_str
        en = ''
    return zh, en


def parse_date(val, year, month_str):
    """將日期數字轉為 YYYY/MM/DD 格式"""
    if val is None or val == '':
        return ''
    try:
        day = int(float(str(val)))
        return f"{year}/{month_str}/{day:02d}"
    except Exception:
        return str(val)


def import_sheet(ws, data, year, log):
    """匯入單一工作表"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # 取得第一個學生列的 meta 資訊
    meta_row = None
    for r in rows[1:]:
        if r[0] and r[1] and r[4] and '=$' not in str(r[4]):
            meta_row = r
            break
    if not meta_row:
        return

    month_zh = str(meta_row[0]).strip()
    cls_name = str(meta_row[1]).strip()
    course   = str(meta_row[2]).strip()
    teacher  = str(meta_row[3]).strip()
    month_str = MONTH_MAP.get(month_zh, '01')
    ym = f"{year}-{month_str}"
    book = COURSE_BOOK_MAP.get(course, course)

    log.append(f"  班級：{cls_name}　課程：{course}　月份：{year}-{month_str}")

    # 確保班級存在
    if cls_name not in data["classes"]:
        data["classes"][cls_name] = {"book": book, "teacher": teacher, "students": []}
        log.append(f"    新建班級 {cls_name} (老師: {teacher})")
    else:
        # 如果班級已存在，也更新一下老師和用書
        data["classes"][cls_name]["book"] = book
        data["classes"][cls_name]["teacher"] = teacher

    data["monthly"].setdefault(cls_name, {})
    data["leveltest"].setdefault(cls_name, [])

    # 解析每個學生（每5列一組）
    i = 1
    while i < len(rows):
        row0 = rows[i]      # 日期列（含學生姓名）
        if i + 4 >= len(rows):
            break
        row1 = rows[i+1]    # 範圍
        row2 = rows[i+2]    # 項目
        row3 = rows[i+3]    # 成績
        row4 = rows[i+4]    # 補考

        # 取學生姓名
        stu_raw = row0[4]
        if not stu_raw or '=$' in str(stu_raw):
            i += 5
            continue

        zh, en = parse_name(stu_raw)

        # 加入學生名單（若不存在）
        if not any(s["zh"] == zh for s in data["classes"][cls_name]["students"]):
            data["classes"][cls_name]["students"].append(
                {"zh": zh, "en": en, "grade": ""}
            )

        # 初始化月成績（清空後重寫，避免重複）
        data["monthly"][cls_name].setdefault(zh, {})
        data["monthly"][cls_name][zh][ym] = []
        added = 0
        records = data["monthly"][cls_name][zh][ym]

        # ── 考試本 col 6-12 ──────────────────────────────────────────────────
        for col in range(6, 13):
            date_val = row0[col]
            if date_val is None:
                continue
            score = row3[col]
            if score is None or score == '':
                continue
            retake = row4[col] if row4[col] is not None else ''
            rng  = str(row1[col]).strip() if row1[col] else ''
            item = str(row2[col]).strip() if row2[col] else ''
            # 過濾公式
            if rng.startswith('=') or item.startswith('='):
                rng = ''; item = ''
            date_str = parse_date(date_val, year, month_str)
            records.append({
                "type": "考試本",
                "date": date_str,
                "range": rng,
                "item": item,
                "score": score,
                "retake": retake,
                "std": 90,
            })
            added += 1

        # ── 口說 col 13-19 ───────────────────────────────────────────────────
        for col in range(13, 20):
            date_val = row0[col]
            if date_val is None:
                continue
            score = row3[col]
            if score is None or score == '':
                continue
            rng = str(row1[col]).strip() if row1[col] else ''
            if rng.startswith('='):
                rng = ''
            
            # 💡 新增：讀取口說的項目
            item = str(row2[col]).strip() if row2[col] else ''
            if item.startswith('='):
                item = ''
                
            date_str = parse_date(date_val, year, month_str)
            records.append({
                "type": "口說",
                "date": date_str,
                "range": rng,
                "item": item if item else '說',
                "score": score,
                "retake": '',
                "std": 90,
            })
            added += 1

        # ── 單字 col 20-26 ───────────────────────────────────────────────────
        for col in range(20, 27):
            date_val = row0[col]
            if date_val is None:
                continue
            score = row3[col]
            if score is None or score == '':
                continue
            retake = row4[col] if row4[col] is not None else ''
            rng = str(row1[col]).strip() if row1[col] else ''
            if rng.startswith('='):
                rng = ''
            
            # 💡 新增：讀取單字的項目
            item = str(row2[col]).strip() if row2[col] else ''
            if item.startswith('='):
                item = ''
                
            date_str = parse_date(date_val, year, month_str)
            records.append({
                "type": "單字",
                "date": date_str,
                "range": rng,
                "item": item if item else 'VOC',
                "score": score,
                "retake": retake,
                "std": 90,
            })
            added += 1

        if added:
            log.append(f"    {zh}({en})：匯入 {added} 筆")

        i += 5


def run_import(xlsx_path, year):
    print(f"開始匯入：{xlsx_path}")
    print(f"年份：{year}")
    print()

    data = load_data()
    wb = load_workbook(xlsx_path, data_only=True)
    log = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.append(f"\n【{sheet_name}】")
        import_sheet(ws, data, year, log)

    save_data(data)

    print("\n".join(log))
    print()
    print(f"✅ 匯入完成！資料已儲存至 {DATA_FILE}")
    print()

    # 統計
    total_cls = len(data["classes"])
    total_stu = sum(len(c["students"]) for c in data["classes"].values())
    print(f"共匯入 {total_cls} 個班級，{total_stu} 位學生")
    input("\n按 Enter 鍵關閉...")


if __name__ == "__main__":
    # 尋找 Excel 檔案
    candidates = list(Path(".").glob("兒美成績輸入表*.xlsx")) + \
                 list(Path(".").glob("*.xlsx"))

    if not candidates:
        print("❌ 找不到 Excel 檔案！")
        print("請將「兒美成績輸入表.xlsx」放在同一個資料夾")
        input("按 Enter 關閉...")
        sys.exit(1)

    xlsx_path = candidates[0]
    print(f"找到檔案：{xlsx_path.name}")

    year_input = input("請輸入成績所屬年份（例如 2026）：").strip()
    if not year_input:
        year_input = "2026"

    run_import(xlsx_path, year_input)
